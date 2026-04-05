#!/usr/bin/env python3
"""
Job Application Agent
7-phase autonomous job search and application system.
Based on: Workflow/job-application-agent.md
Owner: YOUR NAME | your.email@example.com

Usage:
  python agent.py              # Anthropic Claude (requires ANTHROPIC_API_KEY)
  python agent.py --demo       # No API key needed — template/regex mode
  python agent.py --ollama     # Free local LLM via Ollama
  python agent.py --ollama --model mistral  # choose Ollama model
"""

import os
import sys
import json
import re
import argparse
import smtplib
import time
import threading
from email.mime.text import MIMEText
from pathlib import Path
from datetime import date, timedelta, datetime

from rich.console import Console
from rich.panel import Panel
from rich.table import Table

# ─── Config ───────────────────────────────────────────────────────────────────

OWNER_NAME       = "Your Name"  # TODO: replace with your full name
OUTPUT_DIR       = Path("output")
RESOURCES_DIR    = Path("resources")
TODAY            = date.today().strftime("%m/%d/%Y")
MAX_SCRAPE_JOBS  = 50  # cap on total jobs collected in Phase 2 to keep Phase 3 fast

console = Console()


# ─── CLI loading-spinner ──────────────────────────────────────────────────────

class _CliSpinner:
    """Background thread that prints a status message every N seconds
    so the user knows a long-running phase is still alive."""

    _MESSAGES = [
        "Still working — LLM is processing…",
        "Hang tight — fetching results…",
        "Still running — this can take a moment…",
        "Processing in the background…",
        "Almost there — please wait…",
    ]

    def __init__(self, messages=None, interval: int = 20):
        self._msgs = messages or self._MESSAGES
        self._interval = interval
        self._stop = threading.Event()
        self._thread = None

    def start(self):
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        return self

    def stop(self):
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=3)

    def _run(self):
        idx = 0
        while not self._stop.wait(self._interval):
            console.print(f"  [dim]⏳ {self._msgs[idx % len(self._msgs)]}[/dim]")
            idx += 1

    def __enter__(self):
        return self.start()

    def __exit__(self, *_):
        self.stop()


# ─── Module-level counters / state ───────────────────────────────────────────
_last_merge_count: int = 0
_original_latex_source: str = None   # set by _read_resume when LaTeX is detected

# Hardcoded demo job postings used when sample_jobs.json is missing in --demo mode
DEMO_JOBS = [
    {
        "id": "job_001", "title": "IC Design Engineering Intern",
        "company": "NVIDIA", "location": "Santa Clara, CA", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Join NVIDIA's IC design team to work on next-gen GPU silicon.",
        "requirements": ["Verilog", "SPICE", "CMOS", "digital logic", "MATLAB", "Python"],
        "salary_range": "$40–$55/hr", "application_url": "https://nvidia.com/careers/intern-ic",
        "platform": "LinkedIn"
    },
    {
        "id": "job_002", "title": "Photonics Engineering Intern",
        "company": "Lumentum", "location": "San Jose, CA", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Work on photonic integrated circuits and laser component characterization.",
        "requirements": ["Photolithography", "Optical characterization", "MATLAB", "thin film", "Python", "cleanroom"],
        "salary_range": "$35–$45/hr", "application_url": "https://lumentum.com/careers",
        "platform": "Indeed"
    },
    {
        "id": "job_003", "title": "FPGA/Hardware Engineering Intern",
        "company": "Intel", "location": "Hillsboro, OR", "remote": True,
        "posted_date": date.today().isoformat(),
        "description": "Develop and verify FPGA designs for Intel's programmable solutions group.",
        "requirements": ["Verilog", "VHDL", "FPGA", "digital design", "Python", "Linux"],
        "salary_range": "$38–$50/hr", "application_url": "https://intel.com/jobs",
        "platform": "Handshake"
    },
    {
        "id": "job_004", "title": "Semiconductor Process Engineering Intern",
        "company": "Micron Technology", "location": "Boise, ID", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Support semiconductor fabrication process development and yield improvement.",
        "requirements": ["Cleanroom processes", "SPICE", "data analysis", "Python", "MATLAB", "SEM"],
        "salary_range": "$36–$48/hr", "application_url": "https://micron.com/careers",
        "platform": "Glassdoor"
    },
    {
        "id": "job_005", "title": "Mixed-Signal IC Design Intern",
        "company": "Apple", "location": "Cupertino, CA", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Design and simulate mixed-signal circuits for Apple Silicon.",
        "requirements": ["SPICE", "Verilog", "analog design", "CMOS", "Python", "MATLAB"],
        "salary_range": "$45–$60/hr", "application_url": "https://apple.com/jobs",
        "platform": "LinkedIn"
    },
    {
        "id": "job_006", "title": "Hardware Engineering Intern",
        "company": "Microsoft", "location": "Redmond, WA", "remote": True,
        "posted_date": date.today().isoformat(),
        "description": "Contribute to custom silicon and hardware design for Azure infrastructure.",
        "requirements": ["FPGA", "Verilog", "Python", "C++", "digital design", "schematic"],
        "salary_range": "$42–$55/hr", "application_url": "https://microsoft.com/careers",
        "platform": "Indeed"
    },
    {
        "id": "job_007", "title": "Nanoelectronics Research Intern",
        "company": "IBM Research", "location": "Yorktown Heights, NY", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Research novel semiconductor devices and nanoelectronics fabrication methods.",
        "requirements": ["Cleanroom processes", "Photolithography", "device physics", "MATLAB", "Python", "SEM"],
        "salary_range": "$38–$50/hr", "application_url": "https://research.ibm.com/careers",
        "platform": "Handshake"
    },
    {
        "id": "job_008", "title": "EE Hardware Design Intern",
        "company": "Samsung Semiconductors", "location": "San Jose, CA", "remote": False,
        "posted_date": date.today().isoformat(),
        "description": "Support hardware design and verification for Samsung's memory products.",
        "requirements": ["Verilog", "SPICE", "MATLAB", "Python", "digital logic", "PCB design"],
        "salary_range": "$40–$52/hr", "application_url": "https://samsung.com/us/careers",
        "platform": "LinkedIn"
    },
]


# ─── Inference Helpers ────────────────────────────────────────────────────────

def infer_experience_level(job: dict) -> str:
    """Infer experience level from job title and description."""
    text = (job.get("title", "") + " " + job.get("description", "")).lower()
    if any(k in text for k in ["intern", "internship", "co-op", "coop"]):
        return "internship"
    if any(k in text for k in ["entry", "new grad", "graduate", "0-2 years",
                                "junior", "associate"]):
        return "entry-level"
    if any(k in text for k in ["2-5 years", "3+ years", "mid-level"]):
        return "mid-level"
    if any(k in text for k in ["senior", "sr.", "staff", "principal", "lead",
                                "5+ years", "7+ years"]):
        return "senior"
    return "unknown"


def infer_education_required(job: dict) -> str:
    """Infer minimum education requirement from job text."""
    reqs = job.get("requirements", [])
    req_str = " ".join(reqs) if isinstance(reqs, list) else str(reqs)
    text = (job.get("title", "") + " " + job.get("description", "") + " " + req_str).lower()
    if any(k in text for k in ["ph.d", "phd", "doctorate", "doctoral",
                                "doctor of philosophy"]):
        return "phd"
    if any(k in text for k in ["master's", "masters", "m.s.", "ms degree",
                                "meng", "m.eng", "graduate degree"]):
        return "masters"
    if any(k in text for k in ["bachelor's", "bachelors", "bachelor", "b.s.", "bs degree",
                                "b.eng", "undergraduate", "pursuing a degree"]):
        return "bachelors"
    if any(k in text for k in ["associate's", "associates degree", "a.s."]):
        return "associates"
    if any(k in text for k in ["high school diploma", "ged", "no degree required"]):
        return "high_school"
    return "unknown"


def infer_citizenship_required(job: dict) -> str:
    """Infer US citizenship / clearance requirement from job text."""
    reqs = job.get("requirements", [])
    req_str = " ".join(reqs) if isinstance(reqs, list) else str(reqs)
    text = (job.get("title", "") + " " + job.get("description", "") + " " + req_str).lower()
    if any(k in text for k in [
        "us citizen", "u.s. citizen", "united states citizen",
        "security clearance", "secret clearance", "top secret",
        "itar", "us persons only", "must be a us citizen",
        "citizenship required", "due to export control",
    ]):
        return "yes"
    if any(k in text for k in [
        "visa sponsorship available", "open to all work authorizations",
        "no clearance required",
    ]):
        return "no"
    return "unknown"


def deduplicate_jobs(jobs: list) -> list:
    """Merge jobs with same company+title, combining their locations."""
    global _last_merge_count

    def _norm(s: str) -> str:
        return re.sub(r'[^\w\s]', '', s.strip().lower())

    result: list = []
    seen: dict = {}  # key -> index in result

    for job in jobs:
        key = _norm(job.get("company", "")) + "|" + _norm(job.get("title", ""))
        if key not in seen:
            result.append(dict(job))
            seen[key] = len(result) - 1
        else:
            idx = seen[key]
            existing = result[idx]
            # Merge location
            new_loc = job.get("location", "")
            if new_loc and new_loc not in existing.get("location", ""):
                existing["location"] = existing.get("location", "") + f", {new_loc}"
            # Merge URL
            if not existing.get("application_url") and job.get("application_url"):
                existing["application_url"] = job["application_url"]
            # Merge salary
            if not existing.get("salary_range") and job.get("salary_range"):
                existing["salary_range"] = job["salary_range"]
            # Merge remote flag
            if job.get("remote"):
                existing["remote"] = True

    _last_merge_count = len(jobs) - len(result)
    return result


def detect_latex(text: str) -> bool:
    """Return True if text appears to be LaTeX source."""
    if any(text.lstrip().startswith(prefix) for prefix in (
        r"\documentclass", r"\begin{document}"
    )):
        return True
    latex_indicators = [
        r"\section{", r"\subsection{", r"\textbf{",
        r"\cventry", r"\resumeItem", r"\resumeSubheading",
    ]
    if any(ind in text for ind in latex_indicators):
        return True
    if text.count("\\") >= 3:
        return True
    return False


def latex_to_plaintext(latex: str) -> str:
    """Strip LaTeX markup and return clean plain text for LLM parsing."""
    import re as _re
    text = latex

    # Remove preamble (everything before \begin{document})
    doc_start = text.find(r"\begin{document}")
    if doc_start != -1:
        text = text[doc_start + len(r"\begin{document}"):]

    # Remove \end{document}
    text = text.replace(r"\end{document}", "")

    # Remove comments
    text = _re.sub(r"%.*", "", text)

    # Unwrap common formatting commands: \textbf{X} → X
    for cmd in ("textbf", "textit", "emph", "underline", "textsc",
                "textrm", "textsf", "texttt", "small", "large",
                "Large", "huge", "Huge", "normalsize"):
        text = _re.sub(rf"\\{cmd}\{{([^}}]*)\}}", r"\1", text)

    # \href{url}{text} → text
    text = _re.sub(r"\\href\{[^}]*\}\{([^}]*)\}", r"\1", text)

    # \url{...} → (removed)
    text = _re.sub(r"\\url\{[^}]*\}", "", text)

    # Section headings: \section{X} → X (uppercase)
    text = _re.sub(r"\\section\*?\{([^}]*)\}", lambda m: m.group(1).upper(), text)
    text = _re.sub(r"\\subsection\*?\{([^}]*)\}", lambda m: m.group(1).upper(), text)

    # \cventry{dates}{title}{company}{location}{grade}{desc}
    def _cventry(m):
        parts = [p.strip() for p in m.group(1).split("}{")]
        return " | ".join(p for p in parts if p and p != "{}") if parts else ""
    text = _re.sub(r"\\cventry\{(.*?)\}\s*(?=\\|\n|$)", _cventry, text, flags=_re.DOTALL)

    # \resumeSubheading{title}{dates}{company}{location}
    def _subheading(m):
        parts = [p.strip() for p in m.group(1).split("}{")]
        return " | ".join(p for p in parts if p)
    text = _re.sub(r"\\resumeSubheading\{(.*?)\}\s*(?=\\|\n|$)", _subheading, text, flags=_re.DOTALL)

    # \resumeItem{text}
    text = _re.sub(r"\\resumeItem\{([^}]*)\}", r"- \1", text)

    # Remove environments: \begin{X}...\end{X} → just the content
    text = _re.sub(r"\\begin\{[^}]+\}", "", text)
    text = _re.sub(r"\\end\{[^}]+\}", "", text)

    # Remove remaining LaTeX commands with braces: \cmd{text} → text
    text = _re.sub(r"\\[a-zA-Z]+\*?\{([^}]*)\}", r"\1", text)

    # Remove standalone LaTeX commands: \cmd
    text = _re.sub(r"\\[a-zA-Z]+\*?", " ", text)

    # Remove remaining backslash sequences
    text = _re.sub(r"\\.", " ", text)

    # Clean up braces and special chars
    text = text.replace("{", "").replace("}", "")
    text = text.replace("~", " ").replace("&", " | ")

    # Collapse multiple blank lines
    text = _re.sub(r"\n{3,}", "\n\n", text)

    # Strip leading/trailing whitespace per line
    lines = [ln.rstrip() for ln in text.splitlines()]
    text = "\n".join(lines).strip()

    return text


# ─── Provider Abstraction ─────────────────────────────────────────────────────

class BaseProvider:
    """Abstract base for LLM providers."""

    def extract_profile(self, resume_text: str, preferred_titles: list = None) -> dict:
        raise NotImplementedError

    def score_job(self, job: dict, profile: dict) -> dict:
        raise NotImplementedError

    def tailor_resume(self, job: dict, profile: dict, resume_text: str) -> dict:
        raise NotImplementedError

    def generate_cover_letter(self, job: dict, profile: dict) -> str:
        raise NotImplementedError

    def generate_report(self, summary_data: dict) -> str:
        raise NotImplementedError

    def generate_demo_jobs(self, profile: dict, titles: list, location: str) -> list:
        raise NotImplementedError


# ── 1. Anthropic (Claude) Provider ────────────────────────────────────────────

class AnthropicProvider(BaseProvider):
    """Uses Claude Opus 4.6 via the Anthropic API."""

    def __init__(self):
        import anthropic as _anthropic
        self.client = _anthropic.Anthropic()
        self.model  = "claude-opus-4-6"

    def _tool_call(self, tool_def: dict, prompt: str,
                   max_tokens: int = 4096, thinking: bool = False) -> dict:
        kwargs = dict(
            model=self.model, max_tokens=max_tokens,
            tools=[tool_def],
            tool_choice={"type": "tool", "name": tool_def["name"]},
            messages=[{"role": "user", "content": prompt}]
        )
        if thinking:
            kwargs["thinking"] = {"type": "adaptive"}
        resp = self.client.messages.create(**kwargs)
        for block in resp.content:
            if block.type == "tool_use":
                return block.input
        return {}

    def extract_profile(self, resume_text: str, preferred_titles: list = None) -> dict:
        tool = {
            "name": "save_profile",
            "description": "Save the extracted resume profile as structured data.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "name":            {"type": "string"},
                    "email":           {"type": "string"},
                    "linkedin":        {"type": "string"},
                    "location":        {"type": "string"},
                    "target_titles":   {
                        "type": "array", "items": {"type": "string"},
                        "description": (
                            "5–8 specific job titles this candidate is best suited for, "
                            "based on their skills, experience, and education. "
                            "Use industry-standard title variants (e.g. 'IC Design Intern', "
                            "'VLSI Design Engineer', 'Photonics Engineer Intern'). "
                            "Weight heavily toward the candidate's stated preferences."
                        ),
                    },
                    "top_hard_skills": {"type": "array", "items": {"type": "string"}},
                    "top_soft_skills": {"type": "array", "items": {"type": "string"}},
                    "education": {
                        "type": "array",
                        "items": {"type": "object", "properties": {
                            "degree": {"type": "string"}, "institution": {"type": "string"},
                            "year": {"type": "string"}, "gpa": {"type": "string"}
                        }}
                    },
                    "experience": {
                        "type": "array",
                        "items": {"type": "object", "properties": {
                            "title": {"type": "string"}, "company": {"type": "string"},
                            "dates": {"type": "string"},
                            "bullets": {"type": "array", "items": {"type": "string"}}
                        }}
                    },
                    "projects": {
                        "type": "array",
                        "items": {"type": "object", "properties": {
                            "name": {"type": "string"}, "description": {"type": "string"},
                            "skills_used": {"type": "array", "items": {"type": "string"}}
                        }}
                    },
                    "resume_gaps": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["name", "top_hard_skills", "top_soft_skills", "target_titles"]
            }
        }
        pref_hint = ""
        if preferred_titles:
            pref_hint = (
                f"\n\nThe candidate's preferred job titles are: {', '.join(preferred_titles)}. "
                "Use these as a strong signal when generating target_titles — include these "
                "and closely related industry-standard variants that match their background."
            )
        prompt = (
            "Parse this resume. Extract a complete structured profile. "
            "Build a Master Skills Profile: top 10 hard skills and top 5 soft skills "
            "ranked by frequency and recency. Flag any resume gaps.\n"
            "For target_titles: suggest 5–8 specific job titles this candidate is realistically "
            "suited for based on their skills, projects, and experience level. "
            "Use concrete, searchable titles (not generic ones like 'Engineer')."
            f"{pref_hint}\n\n"
            f"Resume:\n{resume_text}"
        )
        return self._tool_call(tool, prompt, thinking=True)

    def score_job(self, job: dict, profile: dict) -> dict:
        tool = {
            "name": "score_job",
            "description": "Score a job posting against the candidate profile.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "job_id":           {"type": "string"},
                    "score":            {"type": "integer", "minimum": 0, "maximum": 100},
                    "matching_skills":  {"type": "array", "items": {"type": "string"}},
                    "missing_skills":   {"type": "array", "items": {"type": "string"}},
                    "reason":           {"type": "string"},
                },
                "required": ["job_id", "score", "matching_skills", "missing_skills", "reason"]
            }
        }
        profile_summary = json.dumps({
            "skills": profile.get("top_hard_skills", []),
            "education": profile.get("education", []),
            "targets": profile.get("target_titles", []),
        })
        prompt = (
            "Score this job (0-100) using: title alignment 25%, skills match 30%, "
            "experience 15%, industry 10%, education 10%, location 10%.\n\n"
            f"Candidate:\n{profile_summary}\n\n"
            f"Job Title: {job.get('title')}\nCompany: {job.get('company')}\n"
            f"Requirements: {', '.join(job.get('requirements', []))}\n"
            f"Description: {job.get('description', '')}\n"
            f"Location: {job.get('location')} (Remote: {job.get('remote', False)})"
        )
        return self._tool_call(tool, prompt, max_tokens=1024)

    def tailor_resume(self, job: dict, profile: dict, resume_text: str) -> dict:
        tool = {
            "name": "tailored_resume",
            "description": "Return tailored resume sections for this specific job.",
            "input_schema": {
                "type": "object",
                "properties": {
                    "skills_reordered":     {"type": "array", "items": {"type": "string"}},
                    "experience_bullets":   {
                        "type": "array",
                        "items": {"type": "object", "properties": {
                            "role": {"type": "string"},
                            "bullets": {"type": "array", "items": {"type": "string"}}
                        }}
                    },
                    "ats_keywords_missing": {"type": "array", "items": {"type": "string"}},
                    "section_order":        {"type": "array", "items": {"type": "string"}},
                },
                "required": ["skills_reordered", "ats_keywords_missing"]
            }
        }
        prompt = (
            f"Tailor this resume for: {job['title']} at {job['company']}.\n"
            "Rules: reorder skills to front-load JD keywords, rephrase bullets naturally — "
            "NEVER fabricate, never change dates/titles/companies. Do NOT include a summary "
            "or objective section.\n"
            "ATS check: flag top JD keywords missing from resume.\n\n"
            f"JD Requirements: {', '.join(job.get('requirements', []))}\n"
            f"JD Description: {job.get('description', '')}\n\n"
            f"Candidate Skills: {', '.join(profile.get('top_hard_skills', []))}\n\n"
            f"Current Resume:\n{resume_text}"
        )
        return self._tool_call(tool, prompt, max_tokens=4096, thinking=True)

    def generate_cover_letter(self, job: dict, profile: dict) -> str:
        resp = self.client.messages.create(
            model=self.model, max_tokens=1024,
            messages=[{"role": "user", "content": (
                f"Write a 3-paragraph cover letter for {OWNER_NAME} applying to "
                f"{job['title']} at {job['company']}.\n"
                "Para 1: Hook + role name. Para 2: Top 2-3 achievements mapped to JD. "
                "Para 3: Enthusiasm + call to action.\n"
                f"Candidate skills: {', '.join(profile.get('top_hard_skills', [])[:5])}\n"
                f"JD requirements: {', '.join(job.get('requirements', [])[:5])}"
            )}]
        )
        return next(b.text for b in resp.content if b.type == "text")

    def generate_report(self, summary_data: dict) -> str:
        resp = self.client.messages.create(
            model=self.model, max_tokens=1024,
            messages=[{"role": "user", "content": (
                "Generate a concise job application run summary.\n\n"
                f"Data:\n{json.dumps(summary_data, indent=2)}\n\n"
                "Include: overall stats, top 3 applied jobs, manual items, "
                "2-3 recommended next steps. Plain text only."
            )}]
        )
        return next(b.text for b in resp.content if b.type == "text")

    def generate_demo_jobs(self, profile: dict, titles: list, location: str) -> list:
        skills = ", ".join(profile.get("top_hard_skills", [])[:5])
        resp = self.client.messages.create(
            model=self.model, max_tokens=4096,
            messages=[{"role": "user", "content": (
                f"Generate 12 realistic internship job postings.\n"
                f"Titles: {', '.join(titles)}\nLocation: {location} or Remote\n"
                f"Key skills: {skills}\n\n"
                "Return a JSON array only (no markdown). Each object: "
                "id, title, company, location, remote (bool), "
                f"posted_date (ISO, last 14 days from {date.today().isoformat()}), "
                "description (2-3 sentences), requirements (array 5-8 strings), "
                "salary_range (string or null), application_url, "
                "platform (LinkedIn|Indeed|Glassdoor|Handshake|Company Site).\n"
                "Focus on IC design, photonics, FPGA, hardware at top EE companies."
            )}]
        )
        raw = next(b.text for b in resp.content if b.type == "text")
        m = re.search(r'\[.*\]', raw, re.DOTALL)
        return json.loads(m.group()) if m else []


# ── 2. Demo Provider (no API, pure Python) ────────────────────────────────────

class DemoProvider(BaseProvider):
    """Template/regex-based provider. Zero cost, zero setup, works offline."""

    # Fallback skill keywords used when config/skill_keywords.yaml is unavailable
    _DEFAULT_KEYWORDS = [
        "verilog", "vhdl", "fpga", "spice", "matlab", "python", "java", "latex",
        "photolithography", "cleanroom", "pld", "cmos", "pcb", "ltspice",
        "onshape", "fusion360", "solidworks", "cad", "linux", "c++",
        "pulsed laser deposition", "thin film", "sem", "afm",
        "digital design", "analog design", "mixed-signal", "rtl", "synthesis",
    ]

    def __init__(self):
        self.SKILL_KEYWORDS = self._load_keywords()

    @staticmethod
    def _load_keywords() -> list:
        """Load and flatten all skill groups from config/skill_keywords.yaml."""
        yaml_path = Path("config/skill_keywords.yaml")
        try:
            import yaml
            with open(yaml_path, encoding="utf-8") as f:
                data = yaml.safe_load(f)
            keywords = []
            for group in data.values():
                if isinstance(group, list):
                    keywords.extend(group)
            return keywords
        except Exception:
            return list(DemoProvider._DEFAULT_KEYWORDS)

    def extract_profile(self, resume_text: str, preferred_titles: list = None) -> dict:
        text_lower = resume_text.lower()

        # Extract email
        email_match = re.search(r'[\w.+-]+@[\w-]+\.[a-zA-Z]{2,}', resume_text)
        email = email_match.group() if email_match else ""

        # Extract LinkedIn
        linkedin_match = re.search(r'linkedin\.com/in/[\w-]+', resume_text, re.I)
        linkedin = linkedin_match.group() if linkedin_match else ""

        # Detect skills present in resume
        found_skills = [s for s in self.SKILL_KEYWORDS if s in text_lower]
        # Capitalize nicely
        skill_display = {
            "verilog": "Verilog", "vhdl": "VHDL", "fpga": "FPGA", "spice": "SPICE",
            "matlab": "MATLAB", "python": "Python", "java": "Java", "latex": "LaTeX",
            "photolithography": "Photolithography", "cleanroom": "Cleanroom Processes",
            "pld": "Pulsed Laser Deposition", "cmos": "CMOS", "pcb": "PCB Design",
            "onshape": "OnShape", "fusion360": "Fusion360", "solidworks": "SolidWorks",
            "cad": "CAD", "linux": "Linux", "c++": "C++",
            "pulsed laser deposition": "Pulsed Laser Deposition",
            "thin film": "Thin Film Deposition", "sem": "SEM",
            "digital design": "Digital Design", "mixed-signal": "Mixed-Signal",
        }
        hard_skills = [skill_display.get(s, s.title()) for s in found_skills[:10]]
        if not hard_skills:
            hard_skills = ["MATLAB", "Python", "Verilog", "FPGA", "SPICE"]

        # Detect degree
        degree = "B.S. Electrical & Computer Engineering"
        institution = "University of Oklahoma"
        if "university of oklahoma" in text_lower or "ou" in text_lower:
            institution = "University of Oklahoma"

        # Extract name (first non-empty line that looks like a name)
        name = OWNER_NAME
        for line in resume_text.splitlines():
            line = line.strip()
            if line and len(line.split()) in (2, 3) and line[0].isupper() and "@" not in line:
                name = line
                break

        # Infer target titles from resume content, then merge with user preferences
        title_map = {
            "fpga": "FPGA/Hardware Engineering Intern",
            "photolithography": "Photonics Engineering Intern",
            "photonics": "Photonics Engineering Intern",
            "spice": "IC Design Engineering Intern",
            "verilog": "IC Design Engineering Intern",
            "vhdl": "VLSI Design Engineering Intern",
            "cmos": "IC Design Engineering Intern",
            "pcb": "Hardware Engineering Intern",
            "mixed-signal": "Mixed-Signal Design Intern",
            "semiconductor": "Semiconductor Process Engineering Intern",
            "thin film": "Thin Film / Materials Engineering Intern",
        }
        inferred = list({title_map[k] for k in title_map if k in text_lower})
        # Merge preferred titles first (user's stated preference takes priority)
        seen = set()
        target_titles = []
        for t in (preferred_titles or []) + inferred:
            if t not in seen:
                seen.add(t)
                target_titles.append(t)
        if not target_titles:
            target_titles = ["IC Design Engineering Intern", "Hardware Engineering Intern"]

        gaps = []
        if "summary" not in text_lower and "objective" not in text_lower:
            gaps.append("Missing professional summary/objective")
        if not re.search(r'\d+%|\d+ students|\d+ projects', resume_text):
            gaps.append("Few quantified achievements — add metrics")

        return {
            "name": name, "email": email, "linkedin": linkedin,
            "location": "Oklahoma, USA",
            "target_titles": target_titles,
            "top_hard_skills": hard_skills,
            "top_soft_skills": ["Teamwork", "Problem-solving", "Communication",
                                 "Attention to detail", "Time management"],
            "education": [{"degree": degree, "institution": institution,
                            "year": "2028", "gpa": ""}],
            "experience": [], "projects": [], "resume_gaps": gaps,
        }

    def score_job(self, job: dict, profile: dict) -> dict:
        skills_lower = {s.lower() for s in profile.get("top_hard_skills", [])}
        reqs = [r.lower() for r in job.get("requirements", [])]

        matched = [r for r in reqs if any(s in r or r in s for s in skills_lower)]
        skills_score = min(30, int(len(matched) / max(len(reqs), 1) * 30))

        # Title alignment
        title_lower = job.get("title", "").lower()
        target_titles_lower = [t.lower() for t in profile.get("target_titles", [])]
        title_score = 20 if any(
            any(word in title_lower for word in t.split()) for t in target_titles_lower
        ) else 10

        # Education: sophomore applying for intern → assume met
        edu_score = 10

        # Location: remote-friendly or Oklahoma
        loc = job.get("location", "").lower()
        remote_ok = job.get("remote", False)
        loc_score = 10 if remote_ok or "oklahoma" in loc or "united states" in loc else 5

        # Industry: EE/semiconductor roles always relevant
        industry_score = 8

        # Experience: student applicant, partial credit
        exp_score = 10

        total = title_score + skills_score + exp_score + industry_score + edu_score + loc_score

        missing = [r.title() for r in reqs if r not in matched and len(r) > 3][:5]
        matched_display = [r.title() for r in matched][:6]

        return {
            "job_id": job.get("id", ""),
            "score": min(total, 100),
            "matching_skills": matched_display,
            "missing_skills": missing,
            "reason": (
                f"Matched {len(matched)}/{len(reqs)} requirements. "
                f"Skills score: {skills_score}/30, Title: {title_score}/25."
            ),
        }

    def tailor_resume(self, job: dict, profile: dict, resume_text: str) -> dict:
        jd_keywords = [r.lower() for r in job.get("requirements", [])]
        skills = profile.get("top_hard_skills", [])
        skills_lower = {s.lower(): s for s in skills}

        # Reorder: JD-matching skills first
        matching = [skills_lower[k] for k in jd_keywords if k in skills_lower]
        other    = [s for s in skills if s not in matching]
        skills_reordered = matching + other

        # ATS gaps: JD keywords not in any skill
        missing_kw = [
            r.title() for r in jd_keywords
            if not any(r in s.lower() or s.lower() in r for s in skills)
        ]

        return {
            "skills_reordered": skills_reordered,
            "experience_bullets": [],
            "ats_keywords_missing": missing_kw[:5],
            "section_order": ["Skills", "Projects", "Experience", "Education"],
        }

    def generate_cover_letter(self, job: dict, profile: dict) -> str:
        skills_str = ", ".join(profile.get("top_hard_skills", [])[:3])
        return (
            f"Dear {job['company']} Hiring Team,\n\n"
            f"I am writing to express my strong interest in the {job['title']} position at "
            f"{job['company']}. As an Electrical & Computer Engineering student at the "
            f"University of Oklahoma with experience in {skills_str}, I am excited to "
            f"contribute to your team.\n\n"
            f"My hands-on lab work in photolithography, cleanroom fabrication, and data "
            f"analysis with MATLAB and Python directly aligns with your requirements. "
            f"I have applied these skills in research projects that mirror the challenges "
            f"your team works on.\n\n"
            f"I would welcome the opportunity to discuss how I can contribute to "
            f"{job['company']}'s goals. Thank you for your consideration.\n\n"
            f"Sincerely,\n{OWNER_NAME}\nsao.sithisack@ou.edu"
        )

    def generate_report(self, summary_data: dict) -> str:
        top3 = summary_data.get("top3_applied", [])
        top3_lines = "\n".join(
            f"  {i+1}. {c} — {t} (score: {s})"
            for i, (c, t, s) in enumerate(top3)
        )
        manual = summary_data.get("manual", 0)
        return (
            f"Run Summary — {date.today().isoformat()}\n\n"
            f"Results:\n"
            f"  • Jobs evaluated:       {summary_data.get('total_found', 0)}\n"
            f"  • Applications sent:    {summary_data.get('applied', 0)}\n"
            f"  • Manual review needed: {manual}\n"
            f"  • Skipped (low match):  {summary_data.get('skipped', 0)}\n\n"
            f"Top Jobs Applied To:\n{top3_lines}\n\n"
            + (f"Manual Review ({manual} item(s)):\n"
               + "\n".join(f"  - {r}" for r in summary_data.get('manual_reasons', []))
               + "\n\n" if manual else "")
            + "Recommended Next Steps:\n"
              "  1. Add quantified metrics to resume bullets (e.g., 'reduced error rate by 20%').\n"
              "  2. Follow up on applied jobs in 7 days via LinkedIn or email.\n"
              "  3. Update skills section with any ATS gaps flagged in tailored resumes.\n"
        )

    def generate_demo_jobs(self, profile: dict, titles: list, location: str) -> list:
        return DEMO_JOBS


# ── 3. Ollama Provider (local free LLM) ───────────────────────────────────────

class OllamaProvider(BaseProvider):
    """Uses a local Ollama model. Free, no API key, requires Ollama installed."""

    OLLAMA_URL = "http://localhost:11434"

    def __init__(self, model: str = "llama3.2"):
        self.model = model
        self._check_ollama()

    def _check_ollama(self):
        import urllib.request as _ur
        import json as _json

        # 1. Check Ollama server is reachable
        try:
            resp = _ur.urlopen(f"{self.OLLAMA_URL}/api/tags", timeout=5)
            data = _json.loads(resp.read().decode())
        except Exception as e:
            raise ConnectionError(
                f"Ollama is not reachable at {self.OLLAMA_URL}.\n"
                f"Start it with:\n"
                f"  ollama serve\n"
                f"(original error: {e})"
            ) from e

        # 2. Check the requested model is pulled locally (fast-fail before LLM call)
        models = data.get("models", [])
        local_bases = {m.get("name", "").split(":")[0] for m in models}
        local_full  = {m.get("name", "") for m in models}
        req_base    = self.model.split(":")[0]

        if self.model not in local_full and req_base not in local_bases:
            available = ", ".join(sorted(local_bases)) or "none"
            raise ValueError(
                f"Model '{self.model}' is not pulled in Ollama.\n"
                f"Available models: {available}\n"
                f"Fix: run  ollama pull {self.model}"
            )

    def _chat(self, prompt: str) -> str:
        """Send prompt to Ollama and return the raw text response.

        json_mode is intentionally NOT used — not all models support
        response_format:json_object and it can cause silent hangs.
        All callers use _parse_json() to extract JSON from plain text.
        """
        try:
            from openai import OpenAI
        except ImportError as exc:
            raise ImportError(
                "openai package is required for Ollama mode. "
                "Run: pip install openai"
            ) from exc

        oc = OpenAI(
            base_url=f"{self.OLLAMA_URL}/v1",
            api_key="ollama",
            timeout=120,  # 2-minute hard timeout per request
        )
        import time
        for attempt in range(5):
            try:
                resp = oc.chat.completions.create(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                )
                return resp.choices[0].message.content or ""
            except Exception as e:
                if "429" in str(e) or "too many concurrent" in str(e).lower():
                    wait = 2 ** attempt
                    console.print(f"  [yellow]⏳ Ollama busy — retrying in {wait}s (attempt {attempt+1}/5)…[/yellow]")
                    time.sleep(wait)
                else:
                    raise
        raise RuntimeError("Ollama rate limit: exceeded 5 retry attempts")

    def _parse_json(self, text: str, fallback: dict) -> dict:
        # Try direct parse, then extract from fences
        for candidate in [text, re.sub(r'^```(?:json)?\s*|\s*```$', '', text, flags=re.M)]:
            try:
                return json.loads(candidate.strip())
            except json.JSONDecodeError:
                pass
        m = re.search(r'\{.*\}', text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                pass
        return fallback

    def extract_profile(self, resume_text: str, preferred_titles: list = None) -> dict:
        pref_hint = ""
        if preferred_titles:
            pref_hint = (
                f"\nThe candidate's preferred roles are: {', '.join(preferred_titles)}. "
                "Weight these heavily when suggesting target_titles; include these and "
                "closely related industry-standard title variants."
            )
        prompt = (
            "Parse this resume and return ONLY a JSON object with these fields:\n"
            "name, email, linkedin, location, "
            "target_titles (array of 5-8 specific searchable job titles this candidate "
            "is best suited for based on their skills and experience level), "
            "top_hard_skills (array of 10), top_soft_skills (array of 5), "
            "education (array of {degree, institution, year, gpa}), "
            "experience (array of {title, company, dates, bullets[]}), "
            "projects (array of {name, description, skills_used[]}), "
            "resume_gaps (array of strings)."
            f"{pref_hint}\n\n"
            f"Resume:\n{resume_text[:3000]}"  # cap to avoid context-limit hangs
        )
        raw = self._chat(prompt)
        return self._parse_json(raw, {
            "name": OWNER_NAME, "email": "", "linkedin": "", "location": "",
            "target_titles": preferred_titles or ["IC Design Intern", "Hardware Engineering Intern"],
            "top_hard_skills": ["MATLAB", "Python", "Verilog", "SPICE", "Photolithography"],
            "top_soft_skills": ["Teamwork", "Problem-solving", "Communication",
                                 "Detail-oriented", "Time management"],
            "education": [], "experience": [], "projects": [], "resume_gaps": []
        })

    def score_job(self, job: dict, profile: dict) -> dict:
        prompt = (
            "Score this job against the candidate (0-100). "
            "Weights: title alignment 25%, skills match 30%, experience 15%, "
            "industry 10%, education 10%, location 10%.\n"
            "Return ONLY a JSON object with: job_id (string), score (int), "
            "matching_skills (array), missing_skills (array), reason (string).\n\n"
            f"Candidate skills: {', '.join(profile.get('top_hard_skills', []))}\n"
            f"Education: {profile.get('education', [{}])[0].get('degree','') if profile.get('education') else ''}\n"
            f"Target titles: {', '.join(profile.get('target_titles', []))}\n\n"
            f"Job: {job.get('title')} at {job.get('company')}\n"
            f"Requirements: {', '.join(job.get('requirements', []))}\n"
            f"Location: {job.get('location')} (Remote: {job.get('remote', False)})"
        )
        raw = self._chat(prompt)
        result = self._parse_json(raw, {
            "job_id": job.get("id", ""), "score": 50,
            "matching_skills": [], "missing_skills": [], "reason": "Scored by Ollama"
        })
        result.setdefault("job_id", job.get("id", ""))
        return result

    def tailor_resume(self, job: dict, profile: dict, resume_text: str) -> dict:
        prompt = (
            f"Tailor this resume for '{job['title']}' at '{job['company']}'.\n"
            "Return ONLY a JSON object with:\n"
            "  skills_reordered (array, JD-matching skills first),\n"
            "  experience_bullets (array of {role, bullets[]}),\n"
            "  ats_keywords_missing (array of JD keywords not in resume),\n"
            "  section_order (array of section names, do NOT include Summary or Objective).\n"
            "Do NOT include a summary or objective field. "
            "NEVER fabricate experience. Only rephrase what exists.\n\n"
            f"JD Requirements: {', '.join(job.get('requirements', []))}\n"
            f"Candidate Skills: {', '.join(profile.get('top_hard_skills', []))}\n\n"
            f"Resume (excerpt):\n{resume_text[:2000]}"
        )
        raw = self._chat(prompt)
        return self._parse_json(raw, {
            "skills_reordered": profile.get("top_hard_skills", []),
            "experience_bullets": [],
            "ats_keywords_missing": [],
            "section_order": ["Skills", "Projects", "Experience", "Education"],
        })

    def generate_cover_letter(self, job: dict, profile: dict) -> str:
        prompt = (
            f"Write a 3-paragraph cover letter for {OWNER_NAME} applying to "
            f"{job['title']} at {job['company']}. "
            "Para 1: hook + role name. Para 2: 2-3 achievements mapped to JD. "
            "Para 3: enthusiasm + CTA. Professional and concise.\n"
            f"Candidate skills: {', '.join(profile.get('top_hard_skills', [])[:5])}"
        )
        return self._chat(prompt)

    def generate_report(self, summary_data: dict) -> str:
        prompt = (
            "Write a concise job application run summary (plain text).\n"
            "Include: overall stats, top 3 jobs, manual items, 2-3 next steps.\n\n"
            f"Data:\n{json.dumps(summary_data, indent=2)}"
        )
        return self._chat(prompt)

    def generate_demo_jobs(self, profile: dict, titles: list, location: str) -> list:
        prompt = (
            f"Generate 10 realistic internship job postings for titles: {', '.join(titles)}. "
            f"Location: {location} or Remote. "
            f"Skills focus: {', '.join(profile.get('top_hard_skills', [])[:5])}.\n"
            "Return ONLY a JSON array. Each item must have: "
            "id, title, company, location, remote (bool), "
            f"posted_date (ISO, within last 14 days from {date.today().isoformat()}), "
            "description (2 sentences), requirements (array 5-8 strings), "
            "salary_range, application_url, platform."
        )
        raw = self._chat(prompt)
        m = re.search(r'\[.*\]', raw, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except json.JSONDecodeError:
                pass
        return DEMO_JOBS  # fallback


# ─── Provider Factory ─────────────────────────────────────────────────────────

def get_provider(args) -> BaseProvider:
    if args.demo:
        console.print("[dim]Mode: Demo (no API key required)[/dim]")
        return DemoProvider()
    if args.ollama:
        console.print(f"[dim]Mode: Ollama local LLM (model: {args.model})[/dim]")
        return OllamaProvider(model=args.model)
    console.print("[dim]Mode: Anthropic Claude Opus 4.6[/dim]")
    return AnthropicProvider()


# ─── Phase 1: Resume Ingestion ────────────────────────────────────────────────

def phase1_ingest_resume(resume_text: str, provider: BaseProvider,
                         preferred_titles: list = None) -> dict:
    console.print("\n[bold cyan]Phase 1 — Resume Ingestion & Profile Extraction[/bold cyan]")
    _t0 = time.time()
    with _CliSpinner(interval=20):
        profile = provider.extract_profile(resume_text, preferred_titles=preferred_titles)
    elapsed = time.time() - _t0
    if profile:
        console.print(f"  ✅ Profile extracted: [bold]{profile.get('name', OWNER_NAME)}[/bold]")
        console.print(f"  📊 Top skills: {', '.join(profile.get('top_hard_skills', [])[:5])}")
        titles = profile.get("target_titles", [])
        if titles:
            console.print(f"  🎯 Target titles: {', '.join(titles[:4])}")
        if profile.get("resume_gaps"):
            console.print(f"  ⚠️  Gaps: {', '.join(profile['resume_gaps'])}")
    console.print(f"  ⏱️  Phase 1 completed in [bold]{elapsed:.1f}s[/bold]")
    return profile


# ─── Job Board Client ─────────────────────────────────────────────────────────

class JobBoardClient:
    """Base class for live job board integrations."""

    def fetch_jobs(self, titles: list, location: str, days: int = 14) -> list:
        """Return a list of job dicts for the given titles and location."""
        raise NotImplementedError


class JobSpyClient(JobBoardClient):
    """Scrapes real job postings from LinkedIn, Indeed, Glassdoor, and ZipRecruiter
    using python-jobspy.  Falls back to an empty list if jobspy is unavailable or
    returns no results, allowing phase2 to trigger the provider fallback.
    """

    def fetch_jobs(self, titles: list, location: str, days: int = 14,
                   max_jobs: int = None) -> list:
        try:
            from jobspy import scrape_jobs
        except ImportError:
            console.print(
                "  [yellow]python-jobspy not installed — "
                "run: pip install python-jobspy[/yellow]"
            )
            return []

        cap = max_jobs or MAX_SCRAPE_JOBS
        # Spread the cap evenly across job titles so we don't over-fetch
        per_title = max(5, cap // max(len(titles), 1))

        all_raw: list = []
        for title in titles:
            if len(all_raw) >= cap:
                break
            try:
                df = scrape_jobs(
                    site_name=["linkedin", "indeed", "glassdoor", "zip_recruiter"],
                    search_term=title,
                    location=location,
                    results_wanted=per_title,
                    hours_old=days * 24,
                    country_indeed="USA",
                )
                all_raw.extend(df.to_dict("records"))
                console.print(f"  📡 '{title}': {len(df)} results scraped")
            except Exception as e:
                console.print(
                    f"  [yellow]JobSpy scrape failed for '{title}': {e}[/yellow]"
                )

        jobs = [self._map(r) for r in all_raw if r.get("job_url")]
        if max_jobs and len(jobs) > max_jobs:
            jobs = jobs[:max_jobs]
            console.print(f"  ✂️  Capped at {max_jobs} jobs (max_scrape_jobs limit)")
        return jobs

    @staticmethod
    def _map(r: dict) -> dict:
        import hashlib

        url   = str(r.get("job_url") or "")
        comp  = str(r.get("company") or "")
        title = str(r.get("title") or "")
        uid   = hashlib.md5(f"{comp}{title}{url}".encode()).hexdigest()[:10]

        # Build a human-readable salary range string
        mn = r.get("min_amount")
        mx = r.get("max_amount")
        iv = r.get("interval") or "yr"
        sal = f"${mn}–${mx}/{iv}" if (mn or mx) else ""

        loc = str(r.get("location") or "")

        # date_posted may be a datetime/date object or a plain string
        pd_raw = r.get("date_posted")
        try:
            posted = pd_raw.isoformat() if hasattr(pd_raw, "isoformat") else str(pd_raw or "")
        except Exception:
            posted = ""

        job = {
            "id":              uid,
            "title":           title,
            "company":         comp,
            "location":        loc,
            "remote":          "remote" in loc.lower() or bool(r.get("is_remote")),
            "posted_date":     posted,
            "description":     str(r.get("description") or ""),
            "requirements":    [],   # LLM extracts requirements from description in phases 3–4
            "salary_range":    sal,
            "application_url": url,
            "platform":        str(r.get("site") or ""),
            "source":          "jobspy",
        }
        job["experience_level"]    = infer_experience_level(job)
        job["education_required"]  = infer_education_required(job)
        job["citizenship_required"] = infer_citizenship_required(job)
        return job


class IndeedClient(JobBoardClient):
    """Legacy stub — kept for backwards compatibility.  JobSpyClient is used instead."""

    def fetch_jobs(self, titles: list, location: str, days: int = 14) -> list:
        return []


class SimplifyJobsScraper:
    """Scrapes the SimplifyJobs Summer 2026 internship board from GitHub."""

    README_URL = (
        "https://raw.githubusercontent.com/SimplifyJobs/"
        "Summer2026-Internships/dev/README.md"
    )

    def fetch_jobs(self, section: str = "hardware") -> list:
        import urllib.request as _ur

        try:
            with _ur.urlopen(self.README_URL, timeout=15) as resp:
                raw = resp.read().decode("utf-8", errors="replace")
        except Exception as e:
            console.print(f"  [yellow]SimplifyJobs fetch failed: {e}[/yellow]")
            return []

        lines = raw.splitlines()

        # Find the Hardware Engineering section header (## line)
        section_start = -1
        for i, line in enumerate(lines):
            if "hardware engineering" in line.lower() and line.strip().startswith("#"):
                section_start = i
                break
        if section_start == -1:
            console.print("  [yellow]SimplifyJobs: Hardware Engineering section not found[/yellow]")
            return []

        # Collect section lines until next ## header
        section_lines = []
        for line in lines[section_start + 1:]:
            if line.strip().startswith("##"):
                break
            section_lines.append(line)

        # Join and parse HTML <tr> rows from the section
        section_text = "\n".join(section_lines)

        # Each <tr>…</tr> block is one job row
        row_blocks = re.findall(r'<tr>(.*?)</tr>', section_text, re.DOTALL | re.I)

        jobs: list = []
        last_company = ""

        for block in row_blocks:
            # Extract <td> cells
            cells = re.findall(r'<td>(.*?)</td>', block, re.DOTALL | re.I)
            if len(cells) < 4:
                continue

            # Cell 0: company  (may be ↳ meaning same company as previous row)
            company_raw = cells[0].strip()
            if company_raw == "↳" or company_raw == "":
                company = last_company
            else:
                company = re.sub(r'<[^>]+>', '', company_raw).strip()
                # Clean HTML entities
                company = company.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
                last_company = company

            # Cell 1: role/title
            title = re.sub(r'<[^>]+>', '', cells[1]).strip()
            # Remove emoji suffixes like 🎓
            title = re.sub(r'[\U00010000-\U0010ffff]', '', title).strip()

            # Cell 2: location (may use <br> for multiple)
            location_raw = cells[2].strip()
            location = re.sub(r'<br\s*/?>', ', ', location_raw, flags=re.I)
            location = re.sub(r'<[^>]+>', '', location).strip()
            location = re.sub(r',\s*,', ',', location).strip(", ")

            # Cell 3: application link
            link_cell = cells[3].strip()

            # Skip closed listings (lock emoji or no href)
            if "\U0001f512" in link_cell or "closed" in link_cell.lower():
                continue

            # Extract first https URL from the link cell
            url_match = re.search(r'href="(https?://[^"]+)"', link_cell)
            if not url_match:
                continue
            url = url_match.group(1)
            # Prefer direct application links over simplify.jobs intermediates when available
            direct = re.findall(r'href="(https?://[^"]+)"', link_cell)
            if direct:
                # Pick non-simplify.jobs link if present
                non_simplify = [u for u in direct if "simplify.jobs" not in u]
                url = non_simplify[0] if non_simplify else direct[0]

            # Cell 4: age/date
            date_cell = cells[4].strip() if len(cells) > 4 else ""
            date_posted = date.today().isoformat()

            if not company or not title:
                continue

            job_stub = {"title": title, "description": "", "requirements": []}
            job = {
                "id":              f"simplify_{abs(hash(company + title + url)) % 100000}",
                "title":           title,
                "company":         company,
                "location":        location,
                "remote":          "remote" in location.lower(),
                "posted_date":     date_posted,
                "description":     f"{title} internship at {company}.",
                "requirements":    [],
                "salary_range":    "",
                "application_url": url,
                "platform":        "SimplifyJobs/GitHub",
                "source":          "simplify",
                "experience_level": "internship",
                "education_required": "unknown",
                "citizenship_required": infer_citizenship_required(job_stub),
            }
            jobs.append(job)

        jobs = deduplicate_jobs(jobs)
        return jobs


# ─── Phase 2: Job Discovery ───────────────────────────────────────────────────

def phase2_discover_jobs(profile: dict, job_titles: list, location: str,
                          provider: BaseProvider,
                          use_simplify: bool = True,
                          max_jobs: int = None) -> list:
    console.print("\n[bold cyan]Phase 2 — Job Discovery & Search[/bold cyan]")
    console.print(f"  🔍 Searching: {', '.join(job_titles)}")
    console.print(f"  📍 Location: {location}")
    cap = max_jobs or MAX_SCRAPE_JOBS
    console.print(f"  🔢 Max jobs cap: {cap}")
    _t0 = time.time()

    sample_file = RESOURCES_DIR / "sample_jobs.json"
    if sample_file.exists():
        with open(sample_file, encoding="utf-8") as f:
            jobs = json.load(f)
        console.print(f"  📂 Loaded {len(jobs)} postings from resources/sample_jobs.json")
        # Apply cap to cached jobs so Phase 3 never scores more than max_jobs
        if cap and len(jobs) > cap:
            jobs = jobs[:cap]
            console.print(f"  ✂️  Capped cached list to {cap} jobs (max_scrape_jobs limit)")
        # Ensure all cached jobs have the inference fields populated
        for job in jobs:
            job.setdefault("experience_level",    infer_experience_level(job))
            job.setdefault("education_required",  infer_education_required(job))
            job.setdefault("citizenship_required", infer_citizenship_required(job))
            job.setdefault("source", "cache")
        console.print(f"  ✅ Returning {len(jobs)} cached jobs")
        console.print(f"  ⏱️  Phase 2 completed in [bold]{time.time() - _t0:.1f}s[/bold]")
        return jobs

    # Try live scrape first; fall back to provider demo generation if needed
    board_client = JobSpyClient()
    with _CliSpinner(
        messages=[
            "Scraping LinkedIn, Indeed, Glassdoor, ZipRecruiter…",
            "Still fetching job listings — network can be slow…",
            "Hang tight — scraping multiple job boards…",
            "Almost done collecting postings…",
            "Deduplicating and filtering results…",
        ],
        interval=20,
    ):
        jobs = board_client.fetch_jobs(job_titles, location, max_jobs=cap)
    if not jobs:
        console.print(
            "  [yellow]⚠️  JobSpy returned 0 results — "
            "falling back to demo job postings.[/yellow]"
        )
        console.print("  🤖 Generating demo job postings...")
        jobs = provider.generate_demo_jobs(profile, job_titles, location)

    # Add inference fields to any jobs that don't have them yet
    for job in jobs:
        job.setdefault("experience_level",    infer_experience_level(job))
        job.setdefault("education_required",  infer_education_required(job))
        job.setdefault("citizenship_required", infer_citizenship_required(job))
        job.setdefault("source", "jobspy")

    # Merge SimplifyJobs listings
    if use_simplify:
        simplify_jobs = SimplifyJobsScraper().fetch_jobs()
        console.print(f"  📋 SimplifyJobs: {len(simplify_jobs)} listings")
        if simplify_jobs:
            simplify_urls = {j["application_url"] for j in simplify_jobs
                             if j.get("application_url")}
            jobs = [j for j in jobs if j.get("application_url") not in simplify_urls]
            jobs = jobs + simplify_jobs

    # Deduplicate by company + title
    before = len(jobs)
    jobs = deduplicate_jobs(jobs)
    after = len(jobs)
    console.print(
        f"  🔀 Deduplication: {before} → {after} jobs "
        f"({before - after} duplicates merged)"
    )

    # Enforce cap after dedup + merge
    if cap and len(jobs) > cap:
        jobs = jobs[:cap]
        console.print(f"  ✂️  Final list capped at {cap} jobs")

    RESOURCES_DIR.mkdir(exist_ok=True)
    with open(sample_file, "w", encoding="utf-8") as f:
        json.dump(jobs, f, indent=2)
    console.print(f"  ✅ {len(jobs)} postings saved to resources/sample_jobs.json")
    console.print(f"  ⏱️  Phase 2 completed in [bold]{time.time() - _t0:.1f}s[/bold]")
    return jobs


# ─── Phase 3: Scoring & Shortlisting ─────────────────────────────────────────

def phase3_score_jobs(jobs: list, profile: dict, provider: BaseProvider,
                       min_score: int = 60,
                       experience_levels=None,
                       education_filter=None,
                       citizenship_filter: str = "all") -> list:
    console.print("\n[bold cyan]Phase 3 — Relevance Scoring & Shortlisting[/bold cyan]")

    filtered = list(jobs)

    # Experience level filter
    if experience_levels and experience_levels != ["all"]:
        filtered = [j for j in filtered
                    if j.get("experience_level", "unknown") in experience_levels]
        console.print(f"  🎯 Experience filter {experience_levels}: {len(filtered)} jobs remain")

    # Education filter (always keep "unknown")
    if education_filter and education_filter != ["all"]:
        filtered = [j for j in filtered if (
            j.get("education_required", "unknown") in education_filter
            or j.get("education_required", "unknown") == "unknown"
        )]
        console.print(f"  🎓 Education filter {education_filter}: {len(filtered)} jobs remain")

    # Citizenship filter
    if citizenship_filter == "exclude_required":
        filtered = [j for j in filtered if j.get("citizenship_required", "unknown") != "yes"]
        console.print(f"  🇺🇸 Citizenship filter (exclude required): {len(filtered)} jobs remain")
    elif citizenship_filter == "only_required":
        filtered = [j for j in filtered if j.get("citizenship_required", "unknown") == "yes"]
        console.print(f"  🇺🇸 Citizenship filter (only required): {len(filtered)} jobs remain")

    console.print(f"  🔢 Scoring {len(filtered)} jobs…")
    _t0_score = time.time()
    scored = []
    for i, job in enumerate(filtered, 1):
        result = provider.score_job(job, profile)
        scored.append({**job, **result})
        if i % 10 == 0 or i == len(filtered):
            elapsed_s = time.time() - _t0_score
            console.print(
                f"  [dim]📊 Scored {i}/{len(filtered)} jobs  "
                f"({elapsed_s:.0f}s elapsed)[/dim]"
            )

    scored = [j for j in scored if j.get("score", 0) >= min_score]
    scored.sort(key=lambda x: x.get("score", 0), reverse=True)

    table = Table(title=f"Job Match Scores (min: {min_score})")
    table.add_column("#",       style="dim",   width=4)
    table.add_column("Company", style="cyan",  width=18)
    table.add_column("Title",   style="white", width=28)
    table.add_column("Score",   style="bold",  width=8)
    table.add_column("Status",  width=20)

    for i, job in enumerate(scored[:12], 1):
        s = job.get("score", 0)
        if s >= 75:
            colour, status = "bold green", "✅ Auto-eligible"
        elif s >= 60:
            colour, status = "yellow", "⚠️  Review needed"
        else:
            colour, status = "red", "❌ Skipped"
        table.add_row(str(i), job.get("company", ""), job.get("title", ""),
                      f"[{colour}]{s}[/{colour}]", status)

    console.print(table)
    console.print(
        f"  ⏱️  Phase 3 completed in [bold]{time.time() - _t0_score:.1f}s[/bold]  "
        f"({len(scored)} jobs passed the {min_score} score threshold)"
    )
    return scored


# ─── Phase 4: Resume Tailoring ────────────────────────────────────────────────

def phase4_tailor_resume(job: dict, profile: dict, resume_text: str,
                          provider: BaseProvider, include_cover_letter: bool = False,
                          section_order: list = None) -> dict:
    tailored = provider.tailor_resume(job, profile, resume_text)
    if section_order:
        tailored["section_order"] = section_order
    if include_cover_letter:
        tailored["cover_letter"] = provider.generate_cover_letter(job, profile)
    return tailored


# ─── Playwright Submitter ─────────────────────────────────────────────────────

class PlaywrightSubmitter:
    """Real application submission via browser automation.

    Requires:  pip install playwright && playwright install chromium
    Activated: python agent.py --real-apply

    Supported boards:
      • boards.greenhouse.io — fills name, email, resume upload, submits
      • All others           — falls back to phase5_simulate_submission()
    """

    def __init__(self, profile: dict):
        self.profile = profile

    def submit(self, job: dict, resume_path: str = "", cover_letter: str = "") -> dict:
        url = job.get("application_url", "")
        if "boards.greenhouse.io" in url:
            return self._submit_greenhouse(job, resume_path)
        return phase5_simulate_submission(job)

    def _submit_greenhouse(self, job: dict, resume_path: str) -> dict:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            console.print(
                "  [yellow]playwright missing — "
                "pip install playwright && playwright install chromium[/yellow]"
            )
            return phase5_simulate_submission(job)

        import random
        url     = job.get("application_url", "")
        profile = self.profile
        name_parts = (profile.get("name") or "").split()
        first  = name_parts[0] if name_parts else ""
        last   = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
        email  = profile.get("email", "")

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()
            try:
                page.goto(url, timeout=30000)
                for sel in ["input[name='first_name']", "input[id='first_name']"]:
                    if page.locator(sel).count():
                        page.fill(sel, first); break
                for sel in ["input[name='last_name']", "input[id='last_name']"]:
                    if page.locator(sel).count():
                        page.fill(sel, last); break
                for sel in ["input[name='email']", "input[id='email']"]:
                    if page.locator(sel).count():
                        page.fill(sel, email); break
                if resume_path and Path(resume_path).exists():
                    for sel in ["input[type='file']", "input[name='resume']"]:
                        if page.locator(sel).count():
                            page.set_input_files(sel, resume_path); break
                for sel in ["button[type='submit']", "input[type='submit']"]:
                    if page.locator(sel).count():
                        page.click(sel)
                        page.wait_for_timeout(2000)
                        break
                return {
                    "status": "Applied",
                    "confirmation": (
                        f"GH-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
                    ),
                }
            except Exception as e:
                console.print(f"  [yellow]Playwright error: {e}[/yellow]")
                return phase5_simulate_submission(job)
            finally:
                browser.close()


# ─── Phase 5: Application Submission (Demo) ───────────────────────────────────

def _load_existing_applications() -> set:
    """Return set of (company_lower, title_lower) already in the current month's tracker."""
    month = datetime.now().strftime("%Y-%m")
    tracker_path = OUTPUT_DIR / f"Job_Applications_Tracker_{month}.xlsx"
    if not tracker_path.exists():
        return set()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(tracker_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        title_col   = headers.index("Job Title") if "Job Title" in headers else None
        company_col = headers.index("Company")   if "Company"   in headers else None
        applied: set = set()
        if title_col is not None and company_col is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                t = row[title_col]
                c = row[company_col]
                if t and c:
                    applied.add((str(c).lower(), str(t).lower()))
        wb.close()
        return applied
    except Exception:
        return set()


def phase5_simulate_submission(job: dict, already_applied: set = None) -> dict:
    if already_applied is None:
        already_applied = set()
    key = (job.get("company", "").lower(), job.get("title", "").lower())
    if key in already_applied:
        console.print("  ⏭️  Already applied — skipped")
        return {"status": "Skipped", "confirmation": "N/A",
                "notes": "Already applied — skipped"}
    import random
    status  = random.choice(["Applied", "Applied", "Applied", "Manual Required"])
    confirm = (
        f"DEMO-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        if status == "Applied" else "N/A"
    )
    return {"status": status, "confirmation": confirm}


# ─── Phase 6: Excel Tracker ───────────────────────────────────────────────────

def phase6_update_tracker(applications: list) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("  [yellow]openpyxl missing — run: pip install openpyxl[/yellow]")
        return None

    console.print("\n[bold cyan]Phase 6 — Excel Tracker[/bold cyan]")

    month        = datetime.now().strftime("%Y-%m")
    tracker_path = OUTPUT_DIR / f"Job_Applications_Tracker_{month}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    headers = [
        "#", "Date Applied", "Job Title", "Company", "Industry",
        "Location", "Job Posting URL", "Company Website", "Application Portal",
        "Match Score", "Resume Version", "Cover Letter Sent",
        "Status", "Confirmation #", "Notes", "Follow-Up Date", "Response Received"
    ]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    status_fills = {
        "Applied":         PatternFill("solid", fgColor="C6EFCE"),
        "Manual Required": PatternFill("solid", fgColor="FFEB9C"),
        "Skipped":         PatternFill("solid", fgColor="FFC7CE"),
        "Error":           PatternFill("solid", fgColor="D9D9D9"),
    }

    for i, app in enumerate(applications, 1):
        applied_str = app.get("date_applied", datetime.now().strftime("%m/%d/%Y"))
        try:
            follow_up = (datetime.strptime(applied_str, "%m/%d/%Y") + timedelta(days=7)).strftime("%m/%d/%Y")
        except ValueError:
            follow_up = ""
        company_slug = app.get("company", "").lower().replace(" ", "")
        ws.append([
            i, applied_str, app.get("title", ""), app.get("company", ""),
            "Technology / Semiconductor", app.get("location", ""),
            app.get("application_url", ""), f"https://www.{company_slug}.com",
            app.get("platform", ""), app.get("score", 0), app.get("resume_version", ""),
            "Yes" if app.get("cover_letter_sent") else "No",
            app.get("status", "Applied"), app.get("confirmation", "N/A"),
            app.get("notes", ""), follow_up, ""
        ])
        fill = status_fills.get(app.get("status", "Applied"), status_fills["Applied"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=col).fill = fill

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Dashboard tab
    ws_d = wb.create_sheet("Dashboard")
    total   = len(applications)
    applied = sum(1 for a in applications if a.get("status") == "Applied")
    manual  = sum(1 for a in applications if a.get("status") == "Manual Required")
    skipped = sum(1 for a in applications if a.get("status") == "Skipped")
    avg_sc  = sum(a.get("score", 0) for a in applications) / max(total, 1)
    for row in [
        ("Metric", "Value"), ("Run Date", date.today().isoformat()),
        ("Total Jobs Evaluated", total), ("Applications Submitted", applied),
        ("Manual Review Required", manual), ("Skipped (Low Match)", skipped),
        ("Average Match Score", f"{avg_sc:.1f}"),
    ]:
        ws_d.append(row)
    ws_d["A1"].font = Font(bold=True)
    ws_d["B1"].font = Font(bold=True)

    wb.save(tracker_path)
    console.print(f"  ✅ Tracker saved → [bold]{tracker_path}[/bold]")
    return tracker_path


# ─── Email Notification ───────────────────────────────────────────────────────

def _send_email_notification(report_text: str, n_applied: int) -> None:
    """Send run-completion email via SMTP_SSL. Silently skipped if any env var is missing."""
    required = ["SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "NOTIFY_EMAIL"]
    missing  = [v for v in required if not os.environ.get(v)]
    if missing:
        console.print(f"  [dim]Email notification skipped (missing env: {', '.join(missing)})[/dim]")
        return
    try:
        host      = os.environ["SMTP_HOST"]
        port      = int(os.environ["SMTP_PORT"])
        user      = os.environ["SMTP_USER"]
        password  = os.environ["SMTP_PASS"]
        recipient = os.environ["NOTIFY_EMAIL"]
        subject   = (
            f"Job Application Run Complete — {date.today().isoformat()} ({n_applied} applied)"
        )
        msg            = MIMEText(report_text)
        msg["Subject"] = subject
        msg["From"]    = user
        msg["To"]      = recipient
        with smtplib.SMTP_SSL(host, port) as smtp:
            smtp.login(user, password)
            smtp.send_message(msg)
        console.print(f"  📧 Notification sent to {recipient}")
    except Exception as e:
        console.print(f"  [yellow]Email notification failed: {e}[/yellow]")


# ─── Phase 7: Run Report ──────────────────────────────────────────────────────

def phase7_run_report(applications: list, tracker_path: Path,
                       provider: BaseProvider) -> str:
    console.print("\n[bold cyan]Phase 7 — End-of-Run Report[/bold cyan]")

    applied_list = [a for a in applications if a.get("status") == "Applied"]
    manual_list  = [a for a in applications if a.get("status") == "Manual Required"]
    skipped_list = [a for a in applications if a.get("status") == "Skipped"]
    top3 = sorted(applied_list, key=lambda x: x.get("score", 0), reverse=True)[:3]

    summary_data = {
        "total_found": len(applications),
        "applied":     len(applied_list),
        "manual":      len(manual_list),
        "skipped":     len(skipped_list),
        "top3_applied": [(a["company"], a["title"], a["score"]) for a in top3],
        "manual_reasons": [a.get("notes", "Form requires manual review") for a in manual_list],
    }

    report_text = provider.generate_report(summary_data)

    report_path = OUTPUT_DIR / f"{datetime.now().strftime('%Y%m%d')}_job-application-run-report.md"
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(f"# Job Application Run Report\n**Date:** {date.today().isoformat()}\n\n")
        f.write(report_text)
        if tracker_path:
            f.write(f"\n\n---\n**Tracker:** `{tracker_path.name}`\n")

    console.print(Panel(report_text, title="[bold]Run Summary[/bold]", border_style="green"))
    console.print(f"  📄 Report saved → [bold]{report_path}[/bold]")
    _send_email_notification(report_text, len(applied_list))
    return report_text


# ─── Startup Checklist ────────────────────────────────────────────────────────

def startup_checklist() -> dict:
    console.print(Panel(
        "[bold white]Job Application Agent[/bold white]\n"
        "7-Phase Autonomous Run  •  Press Enter to accept [defaults]\n"
        "[dim]Modes: default (Claude) | --demo (no API) | --ollama (local LLM)[/dim]",
        border_style="bright_blue",
        title="[bold bright_blue]Startup Checklist[/bold bright_blue]"
    ))

    cfg = {}

    console.print("\n[bold]1. Resume file[/bold]")
    path_str = input("   Path to resume (PDF/DOCX/TXT) [built-in demo profile]: ").strip()
    if path_str:
        cfg["resume_path"] = Path(path_str)
        cfg["resume_text"] = _read_resume(Path(path_str))
        cfg["latex_source"] = _original_latex_source
    else:
        cfg["resume_path"] = None
        cfg["resume_text"] = _build_demo_resume()
        cfg["latex_source"] = None
        console.print("   [dim](i)[/dim]  Using built-in profile from CLAUDE.md")

    console.print("\n[bold]2. Target job titles[/bold] (comma-separated, up to 3)")
    raw = input("   [IC Design Intern, Photonics Engineer Intern, FPGA/Hardware Intern]: ").strip()
    cfg["job_titles"] = (
        [t.strip() for t in raw.split(",")]
        if raw else ["IC Design Intern", "Photonics Engineer Intern", "FPGA/Hardware Intern"]
    )

    console.print("\n[bold]3. Preferred location[/bold]")
    cfg["location"] = input("   [Remote / United States]: ").strip() or "Remote"

    console.print("\n[bold]4. Minimum match score for auto-apply[/bold]")
    raw = input("   [75]: ").strip()
    cfg["threshold"] = int(raw) if raw.isdigit() else 75

    console.print("\n[bold]5. Minimum salary[/bold] (optional)")
    cfg["min_salary"] = input("   [Enter to skip]: ").strip() or None

    console.print("\n[bold]6. Companies to exclude[/bold] (optional)")
    raw = input("   [Enter to skip]: ").strip()
    cfg["blacklist"] = [c.strip() for c in raw.split(",")] if raw else []

    console.print("\n[bold]7. Priority target companies[/bold]")
    raw = input("   [NVIDIA, Apple, Microsoft, Intel, IBM, Micron, Samsung, TSMC]: ").strip()
    cfg["whitelist"] = (
        [c.strip() for c in raw.split(",")]
        if raw else ["NVIDIA", "Apple", "Microsoft", "Intel", "IBM", "Micron", "Samsung", "TSMC"]
    )

    console.print("\n[bold]9. Cover letter preference[/bold]")
    raw = input("   yes / no / only for >=85 [no]: ").strip().lower()
    cfg["cover_letter_mode"] = raw if raw in ("yes", "no", "only for >=85") else "no"

    console.print("\n[bold]10. Max applications per run[/bold]")
    raw = input("   [10]: ").strip()
    cfg["max_apps"] = int(raw) if raw.isdigit() else 10

    return cfg


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _build_demo_resume() -> str:
    return f"""{OWNER_NAME}
Email: sao.sithisack@ou.edu
LinkedIn: www.linkedin.com/in/saoaphisithsithisack
University: University of Oklahoma

OBJECTIVE
Electrical & Computer Engineering sophomore (Spring 2026) seeking summer
internship in IC design, photonics, or hardware engineering.

EDUCATION
University of Oklahoma | B.S. Electrical & Computer Engineering | Expected 2028

TECHNICAL SKILLS
Pulsed Laser Deposition | Photolithography | Cleanroom Processes
MATLAB | Python | Java | LaTeX
CAD: OnShape | Fusion360 | SolidWorks
FPGA | Verilog/VHDL | SPICE Simulation

PROJECTS
Photonics Thin-Film Device
  Deposited thin films using PLD; characterized optical properties with MATLAB.
  Performed photolithography and cleanroom fabrication steps.

IC Prototyping
  Designed and simulated mixed-signal circuits in SPICE.
  Fabricated prototype in university cleanroom.

FPGA Digital Design
  Implemented combinational and sequential logic in Verilog on Xilinx board.

EXPERIENCE
Physics Research Assistant | University of Oklahoma | 2024–Present
  Operated PLD system for thin-film material studies.
  Analyzed experimental data using MATLAB and Python.
  Contributed to lab reports with LaTeX documentation.

STEM Tutor & Mentor (CRLA Level 2) | University of Oklahoma | 2024–Present
  Tutored 20+ students in Physics, Calculus, and EE fundamentals.
  Improved student exam averages by 15% through structured sessions.

INTERESTS
Integrated circuits · Chip design · Photonics · Nanoelectronics · Device physics
"""


def _read_resume(path: Path) -> str:
    global _original_latex_source
    if not path.exists():
        console.print(f"  [yellow]File not found: {path} — using demo resume.[/yellow]")
        return _build_demo_resume()
    suffix = path.suffix.lower()
    if suffix == ".tex":
        raw = path.read_text(encoding="utf-8")
        _original_latex_source = raw
        console.print("  [cyan]LaTeX resume detected — converting to plain text for parsing.[/cyan]")
        return latex_to_plaintext(raw)
    if suffix in (".txt", ".md"):
        raw = path.read_text(encoding="utf-8")
        if detect_latex(raw):
            _original_latex_source = raw
            console.print("  [cyan]LaTeX content detected — converting to plain text for parsing.[/cyan]")
            return latex_to_plaintext(raw)
        return raw
    elif suffix == ".pdf":
        try:
            import pdfplumber
        except ImportError:
            console.print("  [yellow]pdfplumber missing — pip install pdfplumber[/yellow]")
            return _build_demo_resume()
        try:
            with pdfplumber.open(str(path)) as pdf:
                text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            if text.strip():
                return text
            console.print(
                "  [yellow]PDF opened but no text extracted — "
                "the file may be a scanned/image-only PDF.[/yellow]"
            )
            return _build_demo_resume()
        except Exception as e:
            console.print(f"  [yellow]PDF parse error: {e} — using demo resume.[/yellow]")
            return _build_demo_resume()
    elif suffix == ".docx":
        try:
            from docx import Document
            return "\n".join(p.text for p in Document(str(path)).paragraphs if p.text.strip())
        except ImportError:
            console.print("  [yellow]python-docx missing — pip install python-docx[/yellow]")
            return _build_demo_resume()
    else:
        try:
            return path.read_text(encoding="utf-8")
        except Exception:
            console.print(f"  [yellow]Cannot read {path} — using demo resume.[/yellow]")
            return _build_demo_resume()


def remove_summary_section(latex: str) -> str:
    """Strip summary/objective section from a LaTeX resume."""
    import re as _re
    # Remove \section{Summary}, \section{Objective}, \section{Professional Summary}
    # and everything up to the next \section or \end{document}
    pattern = (
        r"\\section\*?\{(?:Summary|Objective|Professional Summary|Career Objective)"
        r"\}.*?(?=\\section|\\end\{document\})"
    )
    return _re.sub(pattern, "", latex, flags=_re.IGNORECASE | _re.DOTALL)


def apply_tailoring_to_latex(latex_source: str, tailored: dict, job: dict) -> str:
    """Inject tailored skills into a LaTeX resume and strip the summary section."""
    import re as _re

    latex = remove_summary_section(latex_source)

    # Replace skills section if provider gave us reordered skills
    skills = tailored.get("skills_reordered")
    if skills:
        skills_line = ", ".join(skills)
        # Try to replace content of a skills section (\section{Skills}...next \section)
        def _replace_skills(m):
            header = m.group(1)  # e.g. \section{Skills}
            rest = m.group(2)    # content before next section
            # Replace itemize/enumerate content or plain text
            new_content = f"\n{skills_line}\n\n"
            # Preserve itemize environment structure if present
            if r"\begin{itemize}" in rest or r"\begin{description}" in rest:
                new_content = (
                    "\n\\begin{itemize}\n"
                    + "".join(f"  \\item {s}\n" for s in skills)
                    + "\\end{itemize}\n\n"
                )
            return header + new_content
        latex = _re.sub(
            r"(\\section\*?\{(?:Skills|Technical Skills|Core Competencies)[^}]*\})"
            r"(.*?)(?=\\section|\\end\{document\})",
            _replace_skills,
            latex,
            flags=_re.IGNORECASE | _re.DOTALL,
        )

    # Append ATS keywords as a comment at the end (invisible in output)
    missing = tailored.get("ats_keywords_missing")
    if missing:
        latex = latex.rstrip()
        if r"\end{document}" in latex:
            latex = latex.replace(
                r"\end{document}",
                f"% ATS gaps: {', '.join(missing)}\n\\end{{document}}"
            )

    return latex


def compile_latex_to_pdf(latex_source: str, output_path: Path) -> bool:
    """Write latex_source to a .tex file and compile it to PDF via pdflatex.

    Returns True on success, False on failure.
    """
    import subprocess
    import tempfile
    import shutil

    if not shutil.which("pdflatex"):
        console.print(
            "  [yellow]pdflatex not found — install TeX Live or MiKTeX to compile PDFs.[/yellow]"
        )
        return False

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        tex_file = tmp / "resume.tex"
        tex_file.write_text(latex_source, encoding="utf-8")
        try:
            result = subprocess.run(
                ["pdflatex", "-interaction=nonstopmode", "resume.tex"],
                cwd=tmpdir,
                capture_output=True,
                text=True,
                timeout=60,
            )
            pdf_file = tmp / "resume.pdf"
            if pdf_file.exists():
                shutil.copy(str(pdf_file), str(output_path))
                return True
            console.print(
                f"  [yellow]pdflatex ran but no PDF produced. "
                f"Stdout: {result.stdout[-500:]}[/yellow]"
            )
            return False
        except subprocess.TimeoutExpired:
            console.print("  [yellow]pdflatex timed out after 60s.[/yellow]")
            return False
        except Exception as e:
            console.print(f"  [yellow]pdflatex error: {e}[/yellow]")
            return False


def _save_tailored_resume(job: dict, tailored: dict, latex_source: str = None) -> str:
    safe = lambda s: re.sub(r"[^a-zA-Z0-9_\-]", "_", s)
    base_name = (
        f"{safe(OWNER_NAME)}_Resume_{safe(job.get('company',''))}"
        f"_{safe(job.get('title',''))}"
    )

    # ── LaTeX / PDF output ────────────────────────────────────────────────────
    if latex_source:
        tailored_latex = apply_tailoring_to_latex(latex_source, tailored, job)
        tex_path = OUTPUT_DIR / (base_name + ".tex")
        tex_path.write_text(tailored_latex, encoding="utf-8")
        pdf_path = OUTPUT_DIR / (base_name + ".pdf")
        compiled = compile_latex_to_pdf(tailored_latex, pdf_path)
        if compiled:
            console.print(f"  [green]PDF saved → {pdf_path.name}[/green]")
            return base_name + ".pdf"
        else:
            console.print(f"  [cyan]LaTeX source saved → {tex_path.name}[/cyan]")
            return base_name + ".tex"

    # ── Plain-text output (original behaviour) ────────────────────────────────
    filename = base_name + ".txt"
    order = (
        tailored.get("section_order")
        or ["Skills", "Projects", "Experience", "Education"]
    )
    with open(OUTPUT_DIR / filename, "w", encoding="utf-8") as f:
        f.write(f"TAILORED RESUME\nRole: {job['title']} @ {job['company']}\n"
                f"Score: {job.get('score','N/A')}\n{'='*60}\n\n")
        for section in order:
            if section == "Skills" and tailored.get("skills_reordered"):
                f.write(f"SKILLS\n{' | '.join(tailored['skills_reordered'])}\n\n")
            elif section == "Experience" and tailored.get("experience_bullets"):
                f.write("EXPERIENCE (tailored)\n")
                for role in tailored["experience_bullets"]:
                    f.write(f"\n{role.get('role','')}\n")
                    for b in role.get("bullets", []):
                        f.write(f"  • {b}\n")
                f.write("\n")
            # Projects and Education sections: not present in tailored dict — skipped
        if tailored.get("ats_keywords_missing"):
            f.write(f"ATS GAPS\n{', '.join(tailored['ats_keywords_missing'])}\n\n")
        if tailored.get("cover_letter"):
            f.write(f"\n{'─'*60}\nCOVER LETTER\n\n{tailored['cover_letter']}\n")
    return filename


# ─── Dashboard Helper ─────────────────────────────────────────────────────────

def _launch_dashboard_and_wait(tracker_path: Path) -> None:
    """Launch the Flask dashboard, open the browser, and wait for Enter."""
    import subprocess
    import webbrowser
    import time

    dashboard_script = Path(__file__).parent / "dashboard" / "app.py"
    if not dashboard_script.exists():
        console.print("  [yellow]dashboard/app.py not found — skipping dashboard.[/yellow]")
        return

    try:
        proc = subprocess.Popen([sys.executable, str(dashboard_script)])
        time.sleep(1.5)  # let Flask start
        webbrowser.open("http://localhost:5000")
        console.print(Panel(
            "Dashboard running at [bold]http://localhost:5000[/bold]\n\n"
            "• Review scored jobs and approve Manual Required rows.\n"
            "• Approved rows will be submitted in phases 4-7.\n"
            "• Press [bold]Enter[/bold] when ready to continue.",
            title="[bold cyan]Web Dashboard[/bold cyan]",
            border_style="cyan",
        ))
        input()
        proc.terminate()
    except Exception as e:
        console.print(f"  [yellow]Dashboard launch failed: {e}[/yellow]")


# ─── Main Orchestrator ────────────────────────────────────────────────────────

def run_agent(config: dict, provider: BaseProvider):
    OUTPUT_DIR.mkdir(exist_ok=True)
    RESOURCES_DIR.mkdir(exist_ok=True)

    console.print("\n" + "─" * 64)
    console.print("[bold green]🤖  Job Application Agent — Run Starting[/bold green]")
    console.print("─" * 64)

    profile = phase1_ingest_resume(
        config["resume_text"], provider,
        preferred_titles=config.get("job_titles"),
    )
    if not profile:
        console.print("[red]Phase 1 failed — cannot parse resume.[/red]")
        return

    jobs = phase2_discover_jobs(
        profile, config["job_titles"], config["location"], provider,
        use_simplify=config.get("use_simplify", True),
        max_jobs=config.get("max_scrape_jobs", MAX_SCRAPE_JOBS),
    )
    if config.get("blacklist"):
        bl = {c.lower() for c in config["blacklist"]}
        jobs = [j for j in jobs if j.get("company", "").lower() not in bl]

    threshold = config.get("threshold", 75)
    scored = phase3_score_jobs(
        jobs, profile, provider, min_score=60,
        experience_levels=config.get("experience_levels"),
        education_filter=config.get("education_filter"),
        citizenship_filter=config.get("citizenship_filter", "all"),
    )

    auto_eligible = [j for j in scored if j.get("score", 0) >= threshold]
    review_needed = [j for j in scored if 60 <= j.get("score", 0) < threshold]

    console.print(
        f"\n  📋 [bold]{len(auto_eligible)}[/bold] auto-eligible (≥{threshold})  |  "
        f"[yellow]{len(review_needed)}[/yellow] for review  |  "
        f"[red]{len(scored) - len(auto_eligible) - len(review_needed)}[/red] skipped"
    )

    console.print("\n[bold]Top auto-eligible jobs:[/bold]")
    for j in auto_eligible[:5]:
        console.print(
            f"  [{j['score']}] {j['company']} — {j['title']}  "
            f"[dim]{j.get('location','?')} · {j.get('platform','')}[/dim]"
        )

    if config.get("dashboard"):
        # Write a preliminary tracker so the dashboard has data to display
        preliminary_apps = [
            {**j, "date_applied": datetime.now().strftime("%m/%d/%Y"),
             "resume_version": "", "cover_letter_sent": False,
             "status": "Auto-eligible" if j.get("score", 0) >= config.get("threshold", 75)
                       else "Manual Required",
             "confirmation": "N/A", "notes": ""}
            for j in scored
        ]
        phase6_update_tracker(preliminary_apps)
        _launch_dashboard_and_wait(OUTPUT_DIR / f"Job_Applications_Tracker_{datetime.now().strftime('%Y-%m')}.xlsx")

    proceed = input("\nProceed with submissions? [Y/n]: ").strip().lower()
    if proceed == "n":
        console.print("[yellow]Run cancelled.[/yellow]")
        return

    to_process      = auto_eligible[:config.get("max_apps", 10)]
    already_applied = _load_existing_applications()
    applications    = []
    submitter       = PlaywrightSubmitter(profile) if config.get("real_apply") else None

    for i, job in enumerate(to_process, 1):
        console.print(f"\n[bold]({i}/{len(to_process)}) {job['title']} @ {job['company']}[/bold]  score={job['score']}")

        include_cl = (
            config.get("cover_letter_mode") == "yes"
            or (config.get("cover_letter_mode") == "only for >=85"
                and job.get("score", 0) >= 85)
        )

        console.print("  ✏️  Tailoring resume...")
        tailored = phase4_tailor_resume(job, profile, config["resume_text"],
                                         provider, include_cl,
                                         section_order=config.get("section_order"))

        if tailored.get("ats_keywords_missing"):
            console.print(
                f"  ⚠️  ATS gaps: [yellow]{', '.join(tailored['ats_keywords_missing'][:4])}[/yellow]"
            )

        resume_file = _save_tailored_resume(job, tailored, config.get("latex_source"))
        console.print(f"  💾 Resume → output/{resume_file}")

        if submitter:
            console.print("  🚀 Submitting via Playwright...")
            result = submitter.submit(job, str(OUTPUT_DIR / resume_file))
        else:
            console.print("  🚀 Submitting (demo mode)...")
            result = phase5_simulate_submission(job, already_applied)
        icon   = "✅" if result["status"] == "Applied" else "⚠️"
        console.print(f"  {icon} {result['status']}  •  Confirmation: {result['confirmation']}")

        applications.append({
            **job,
            "date_applied":      datetime.now().strftime("%m/%d/%Y"),
            "resume_version":    resume_file,
            "cover_letter_sent": bool(tailored.get("cover_letter")),
            "status":            result["status"],
            "confirmation":      result["confirmation"],
            "notes": (
                "ATS gaps: " + ", ".join(tailored.get("ats_keywords_missing", [])[:2])
                if tailored.get("ats_keywords_missing") else ""
            ),
        })

    for job in scored:
        if job not in to_process:
            applications.append({
                **job,
                "date_applied": datetime.now().strftime("%m/%d/%Y"),
                "resume_version": "", "cover_letter_sent": False,
                "status": "Skipped", "confirmation": "N/A",
                "notes": f"Score {job.get('score',0)} below threshold or max_apps reached",
            })

    tracker_path = phase6_update_tracker(applications)
    phase7_run_report(applications, tracker_path, provider)

    console.print("\n[bold green]✅  Agent run complete![/bold green]")
    if tracker_path:
        console.print(f"   📊 Tracker → [bold]{tracker_path}[/bold]")
    console.print("\n[dim]Next: review Manual Required rows, update 'Response Received' as replies arrive.[/dim]")


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Job Application Agent — autonomous 7-phase job search system",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python agent.py                   # Claude (needs ANTHROPIC_API_KEY)\n"
            "  python agent.py --demo            # No API key needed\n"
            "  python agent.py --ollama          # Local Ollama (free)\n"
            "  python agent.py --ollama --model mistral\n"
        )
    )
    parser.add_argument("--demo",   action="store_true",
                        help="Run without any LLM — template/regex mode, zero cost")
    parser.add_argument("--ollama", action="store_true",
                        help="Use local Ollama LLM (free, requires ollama.com)")
    parser.add_argument("--model",  default="llama3.2",
                        help="Ollama model name (default: llama3.2)")
    parser.add_argument("--section-order", default=None,
                        help="Comma-separated resume section order "
                             "(e.g. Summary,Skills,Experience,Projects,Education)")
    parser.add_argument("--real-apply", action="store_true",
                        help="Use Playwright for real form submission (Greenhouse boards)")
    parser.add_argument("--dashboard", action="store_true",
                        help="Launch Flask dashboard after scoring for manual review")
    parser.add_argument(
        "--experience", default="internship,entry-level",
        help="Comma-separated experience levels to include "
             "(internship, entry-level, mid-level, senior, unknown, all). "
             "Default: internship,entry-level",
    )
    parser.add_argument(
        "--education", default="bachelors,masters,unknown",
        help="Comma-separated education levels to include "
             "(high_school, associates, bachelors, masters, phd, unknown, all). "
             "Default: bachelors,masters,unknown",
    )
    parser.add_argument(
        "--citizenship", default="all",
        choices=["all", "exclude_required", "only_required"],
        help="Citizenship filter. Default: all",
    )
    parser.add_argument(
        "--no-simplify", action="store_true",
        help="Skip SimplifyJobs/GitHub internship listings",
    )
    args = parser.parse_args()

    # API key guard — only needed for default Anthropic mode
    if not args.demo and not args.ollama and not os.environ.get("ANTHROPIC_API_KEY"):
        console.print(
            "[red]Error: ANTHROPIC_API_KEY not set.[/red]\n\n"
            "Options:\n"
            "  1. Set key:  [bold]set ANTHROPIC_API_KEY=sk-ant-...[/bold]  (Windows)\n"
            "  2. No key?   [bold]python agent.py --demo[/bold]\n"
            "  3. Free LLM: [bold]python agent.py --ollama[/bold]  (install ollama.com first)"
        )
        sys.exit(1)

    provider = get_provider(args)
    config   = startup_checklist()
    config["section_order"] = (
        [s.strip() for s in args.section_order.split(",")]
        if args.section_order else None
    )
    config["real_apply"]        = args.real_apply
    config["dashboard"]         = args.dashboard
    config["use_simplify"]      = not args.no_simplify
    config["experience_levels"] = [x.strip() for x in args.experience.split(",")]
    config["education_filter"]  = [x.strip() for x in args.education.split(",")]
    config["citizenship_filter"] = args.citizenship
    run_agent(config, provider)
