"""
Flask Web UI Dashboard for Job Application Agent.

Single-page table of all jobs from the current month's tracker, with
one-click approval for Manual Required rows.

Run standalone:
    python dashboard/app.py

Or launch automatically via:
    python agent.py --demo --dashboard
"""

import os
import secrets
import sys
from pathlib import Path
from datetime import datetime

try:
    from flask import Flask, abort, render_template_string, redirect, request, session, url_for, jsonify
except ImportError:
    print("Flask not installed. Run: pip install flask")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

OUTPUT_DIR = Path(__file__).parent.parent / "output"

app = Flask(__name__)
# Secret key required for the signed session cookie (used for CSRF tokens).
app.secret_key = os.environ.get("DASHBOARD_SECRET") or secrets.token_urlsafe(32)
# Shared secret required on every request. Set DASHBOARD_TOKEN to enable auth.
_DASHBOARD_TOKEN = os.environ.get("DASHBOARD_TOKEN", "")


def _csrf_token() -> str:
    tok = session.get("_csrf")
    if not tok:
        tok = secrets.token_urlsafe(24)
        session["_csrf"] = tok
    return tok


def _require_dashboard_auth() -> None:
    """Block unauthenticated access if a token is configured.

    Without DASHBOARD_TOKEN set the dashboard runs open *only* on loopback
    (the bind below enforces that). With DASHBOARD_TOKEN set, callers must
    present it via either a `?token=` query arg or the `X-Dashboard-Token`
    header — checked with constant-time comparison.
    """
    if not _DASHBOARD_TOKEN:
        return
    provided = request.args.get("token") or request.headers.get("X-Dashboard-Token") or ""
    if not secrets.compare_digest(provided, _DASHBOARD_TOKEN):
        abort(401)


@app.before_request
def _enforce_dashboard_auth():
    _require_dashboard_auth()

_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Job Application Dashboard</title>
<style>
  body  { font-family: sans-serif; margin: 2rem; background: #f5f5f5; color: #222; }
  h1    { color: #1F4E79; margin-bottom: 0.25rem; }
  .meta { color: #555; font-size: 0.9rem; margin-bottom: 1.5rem; }
  table { border-collapse: collapse; width: 100%; background: white;
          box-shadow: 0 1px 4px rgba(0,0,0,.12); }
  th    { background: #1F4E79; color: white; padding: 9px 14px;
          text-align: left; white-space: nowrap; }
  td    { padding: 8px 14px; border-bottom: 1px solid #e0e0e0; vertical-align: middle; }
  tr.Applied  td { background: #C6EFCE; }
  tr.Manual   td { background: #FFEB9C; }
  tr.Skipped  td { background: #FFC7CE; }
  tr.Approved td { background: #d0e8ff; }
  a     { color: #1a5fa8; }
  .approve-btn {
    background: #28a745; color: white; border: none; padding: 4px 12px;
    border-radius: 4px; cursor: pointer; font-size: 0.85rem;
  }
  .approve-btn:hover { background: #1e7e34; }
  .badge {
    display: inline-block; padding: 2px 8px; border-radius: 10px;
    font-size: 0.8rem; font-weight: bold;
  }
  .badge-applied  { background: #28a745; color: white; }
  .badge-manual   { background: #ffc107; color: #333; }
  .badge-skipped  { background: #dc3545; color: white; }
  .badge-approved { background: #17a2b8; color: white; }
</style>
</head>
<body>
<h1>Job Application Dashboard</h1>
<p class="meta">
  Tracker: <code>{{ tracker_name }}</code> &nbsp;|&nbsp;
  {{ jobs|length }} entries &nbsp;|&nbsp;
  Refreshed: {{ now }}
</p>
<table>
<thead>
<tr>
  <th>#</th><th>Date</th><th>Company</th><th>Title</th>
  <th>Score</th><th>Status</th><th>URL</th><th>Action</th>
</tr>
</thead>
<tbody>
{% for j in jobs %}
<tr class="{{ j.row_class }}">
  <td>{{ j.num }}</td>
  <td>{{ j.date }}</td>
  <td><strong>{{ j.company }}</strong></td>
  <td>{{ j.title }}</td>
  <td>{{ j.score }}</td>
  <td>
    {% if j.status == 'Applied' %}
      <span class="badge badge-applied">Applied</span>
    {% elif j.status == 'Manual Required' %}
      <span class="badge badge-manual">Manual Review</span>
    {% elif j.status == 'Approved' %}
      <span class="badge badge-approved">Approved</span>
    {% else %}
      <span class="badge badge-skipped">{{ j.status }}</span>
    {% endif %}
  </td>
  <td>{% if j.url %}<a href="{{ j.url }}" target="_blank">Open</a>{% endif %}</td>
  <td>
    {% if j.status == 'Manual Required' %}
    <form method="post" action="/approve/{{ j.num }}" style="display:inline">
      <input type="hidden" name="csrf_token" value="{{ csrf_token }}">
      <button class="approve-btn" type="submit">Approve</button>
    </form>
    {% endif %}
  </td>
</tr>
{% endfor %}
</tbody>
</table>
</body>
</html>"""


def _tracker_path() -> Path:
    month = datetime.now().strftime("%Y-%m")
    return OUTPUT_DIR / f"Job_Applications_Tracker_{month}.xlsx"


def _load_jobs():
    path = _tracker_path()
    if not path.exists():
        return [], path.name
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    jobs = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue
        rd = dict(zip(headers, row))
        status = rd.get("Status", "") or ""
        row_class = {
            "Applied":         "Applied",
            "Manual Required": "Manual",
            "Approved":        "Approved",
        }.get(status, "Skipped")
        jobs.append({
            "num":       i,
            "date":      rd.get("Date Applied", ""),
            "company":   rd.get("Company", ""),
            "title":     rd.get("Job Title", ""),
            "score":     rd.get("Match Score", ""),
            "status":    status,
            "url":       rd.get("Job Posting URL", ""),
            "row_class": row_class,
        })
    wb.close()
    return jobs, path.name


@app.route("/")
def index():
    jobs, tracker_name = _load_jobs()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    return render_template_string(
        _TEMPLATE,
        jobs=jobs,
        tracker_name=tracker_name,
        now=now,
        csrf_token=_csrf_token(),
    )


@app.route("/approve/<int:job_id>", methods=["POST"])
def approve(job_id):
    """Mark a Manual Required row as Approved in the tracker."""
    submitted = request.form.get("csrf_token", "")
    expected = session.get("_csrf", "")
    if not expected or not secrets.compare_digest(submitted, expected):
        abort(400, "Invalid CSRF token")
    if job_id < 1:
        abort(400, "Invalid row id")

    path = _tracker_path()
    if not path.exists():
        return jsonify({"error": "Tracker not found"}), 404
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        status_col = headers.index("Status") + 1 if "Status" in headers else None
        target_row = job_id + 1  # row 1 = header
        if status_col and 2 <= target_row <= ws.max_row:
            current = ws.cell(row=target_row, column=status_col).value
            # Only flip rows that were waiting on manual review; refuse otherwise
            # so a stale browser tab can't clobber an Applied row.
            if current != "Manual Required":
                abort(409, "Row is not in Manual Required state")
            ws.cell(row=target_row, column=status_col).value = "Approved"
            blue_fill = PatternFill("solid", fgColor="BDD7EE")
            for col in range(1, len(headers) + 1):
                ws.cell(row=target_row, column=col).fill = blue_fill
            wb.save(path)
    finally:
        wb.close()
    return redirect(url_for("index"))


if __name__ == "__main__":
    bind_host = os.environ.get("DASHBOARD_HOST", "127.0.0.1")
    bind_port = int(os.environ.get("DASHBOARD_PORT", "5000"))
    print(f"Dashboard reading from: {OUTPUT_DIR}")
    if not _DASHBOARD_TOKEN and bind_host not in ("127.0.0.1", "localhost", "::1"):
        print(
            "Refusing to bind to non-loopback host without DASHBOARD_TOKEN set.",
            file=sys.stderr,
        )
        sys.exit(2)
    app.run(host=bind_host, port=bind_port, debug=False)
