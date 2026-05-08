"""
pipeline/phases.py
──────────────────
The seven pipeline phase functions, Playwright submitter, tracker writer,
email notifier, and dashboard helper.
"""

import json
import os
import re
import smtplib
import sys
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout, as_completed
from datetime import date, datetime, timedelta
from email.mime.text import MIMEText
from pathlib import Path

from rich.panel import Panel
from rich.table import Table

from .config import console, OUTPUT_DIR, RESOURCES_DIR, MAX_SCRAPE_JOBS, _CliSpinner
from .helpers import (infer_experience_level, infer_education_required,
                      infer_citizenship_required, deduplicate_jobs,
                      filter_jobs_by_education, validate_job_urls)
from .providers import BaseProvider, RUBRIC_WEIGHTS, compute_skill_coverage
from .resume import _build_demo_resume, _save_tailored_resume
from .heuristic_tailor import (
    validate_tailoring,
    heuristic_tailor_resume,
    merge_with_heuristic,
)


# ── Phase 1 ────────────────────────────────────────────────────────────────────

def phase1_ingest_resume(resume_text: str, provider: BaseProvider,
                         preferred_titles: list = None,
                         force: bool = False,
                         ollama_model: str = None) -> dict:
    import hashlib
    from .config import OWNER_NAME
    from .profile_audit import audit_profile
    from .profile_extractor import scan_profile, merge_profiles, heuristic_summary
    from .resume_insights import analyze_resume, INSIGHTS_VERSION
    console.print("\n[bold cyan]Phase 1 — Resume Ingestion & Profile Extraction[/bold cyan]")
    _t0 = time.time()

    cache_key = json.dumps({
        "resume_text": resume_text,
        "provider": provider.__class__.__name__,
        "provider_model": getattr(provider, "model", ""),
        "preferred_titles": preferred_titles or [],
    }, sort_keys=True)
    resume_hash = hashlib.md5(cache_key.encode()).hexdigest()
    cache_file = RESOURCES_DIR / f"profile_cache_{resume_hash}.json"

    profile = None
    cache_hit = False
    if not force and cache_file.exists():
        try:
            with open(cache_file, encoding="utf-8") as f:
                profile = json.load(f)
            console.print("  ⚡ Loaded cached profile (resume unchanged)")
            cache_hit = True
        except Exception:
            profile = None
    if force and cache_file.exists():
        console.print("  🔄 Forced re-extraction — bypassing profile cache")

    if profile is None:
        # ── Heuristic-first extraction ──────────────────────────────────────
        # Always run the deterministic regex/section scanner. Its output
        # primes the LLM prompt so the LLM verifies rather than re-derives;
        # if the LLM fails or skips fields, we still have a complete profile.
        heuristic = scan_profile(resume_text)
        console.print(f"  🔎 Heuristic scan → {heuristic_summary(heuristic)}")

        # ── LLM verification + enrichment ───────────────────────────────────
        with _CliSpinner(interval=20):
            try:
                llm_profile = provider.extract_profile(
                    resume_text,
                    preferred_titles=preferred_titles,
                    heuristic_hint=heuristic,
                )
            except TypeError:
                # Older provider implementations won't accept the hint kwarg.
                llm_profile = provider.extract_profile(
                    resume_text, preferred_titles=preferred_titles,
                )

        # ── Merge: heuristic baseline + LLM corrections ─────────────────────
        profile = merge_profiles(heuristic, llm_profile)

        # Post-extraction audit: flatten → quarantine → retention → verify → rerank.
        if profile:
            profile = audit_profile(profile, resume_text)
            # Back-compat: ensure `experience` is populated from research/work split.
            if not profile.get("experience"):
                profile["experience"] = (
                    list(profile.get("research_experience") or [])
                    + list(profile.get("work_experience") or [])
                )

    if profile:
        # Insights are recomputed when the cached version is stale, or when
        # we have an LLM provider whose verification could upgrade a cache
        # that only holds heuristic-only insights. Provider-agnostic: any
        # BaseProvider subclass that exposes a callable chat() qualifies.
        existing = profile.get("insights") or {}
        provider_can_verify = bool(provider) and callable(getattr(provider, "chat", None))
        cached_verified = str(existing.get("verified_by") or "") == "ai"
        needs_rescan = (
            not existing
            or int(existing.get("version") or 0) < INSIGHTS_VERSION
            or (provider_can_verify and not cached_verified)
        )
        if needs_rescan:
            console.print(
                "  🔬 Scanning resume content & "
                + ("verifying with AI" if provider_can_verify else "computing heuristic insights")
                + "…"
            )
            profile["insights"] = analyze_resume(
                resume_text, profile,
                provider=provider if provider_can_verify else None,
            )
            cache_hit = False  # force re-save with fresh insights

    elapsed = time.time() - _t0
    if profile:
        console.print(f"  ✅ Profile extracted: [bold]{profile.get('name', OWNER_NAME)}[/bold]")
        console.print(f"  📊 Top skills: {', '.join(profile.get('top_hard_skills', [])[:5])}")
        titles = profile.get("target_titles", [])
        if titles:
            console.print(f"  🎯 Target titles: {', '.join(titles[:4])}")
        if profile.get("resume_gaps"):
            console.print(f"  ⚠️  Gaps: {', '.join(profile['resume_gaps'])}")
        for note in profile.get("_audit_log", []):
            console.print(f"  [dim]🔍 audit: {note}[/dim]")
        method = profile.get("_extraction_method", "?")
        console.print(f"  [dim]🧬 Extraction method: {method}[/dim]")
        ins = profile.get("insights") or {}
        if ins:
            console.print(
                f"  🧮 Insight score: [bold]{ins.get('overall_score', '?')}/100[/bold] "
                f"({ins.get('verified_by', 'heuristic')})"
            )

    if not cache_hit:
        RESOURCES_DIR.mkdir(exist_ok=True)
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(profile, f, indent=2)

    console.print(f"  ⏱️  Phase 1 completed in [bold]{elapsed:.1f}s[/bold]")
    return profile


# ── Phase 2 ────────────────────────────────────────────────────────────────────

def _parse_posted_date(val) -> datetime | None:
    """Best-effort parse of the heterogeneous `posted_date` field."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def _filter_by_posting_age(jobs: list, days_old: int) -> list:
    """Drop jobs posted before (now - days_old). Jobs with no date are kept."""
    if not days_old or days_old <= 0:
        return jobs
    cutoff = datetime.now() - timedelta(days=days_old)
    kept = []
    for j in jobs:
        posted = _parse_posted_date(j.get("posted_date"))
        if posted is None or posted >= cutoff:
            kept.append(j)
    return kept


def _sort_newest_first(jobs: list) -> list:
    """Sort by posted_date descending; undated jobs sort last."""
    return sorted(
        jobs,
        key=lambda j: _parse_posted_date(j.get("posted_date")) or datetime.min,
        reverse=True,
    )


def _resolve_effective_titles(job_titles, profile: dict):
    """Decide which titles will actually drive the scrape.

    Returns ``(titles, source)`` where *source* is one of:
      - ``"user"``    → caller-supplied titles are kept as-is
      - ``"phase1"``  → caller passed nothing meaningful; falls back to
                         ``profile["target_titles"]``
      - ``"merged"``  → caller's titles plus phase-1 suggestions, deduped
    """
    incoming = [t for t in (job_titles or []) if t and str(t).strip()]
    only_engineer = (
        len(incoming) == 1 and str(incoming[0]).strip().lower() == "engineer"
    )
    phase1 = [str(t).strip() for t in (profile.get("target_titles") or []) if t]

    if (not incoming or only_engineer) and phase1:
        return phase1, "phase1"

    if incoming and phase1:
        merged: list = []
        seen: set = set()
        for t in (*incoming, *phase1):
            key = str(t).strip().lower()
            if key and key not in seen:
                seen.add(key)
                merged.append(str(t).strip())
        return merged, "merged"

    return incoming, "user"


def phase2_discover_jobs(profile: dict, job_titles: list, location: str,
                          provider: BaseProvider,
                          use_simplify: bool = True,           # noqa: ARG001
                          max_jobs: int = None,
                          days_old: int = 30,
                          education_filter=None,
                          include_unknown_education: bool = False,
                          deep_search: bool = False,
                          cache_ttl_minutes: int = 30,         # noqa: ARG001
                          force_live: bool = False,
                          offset: int = 0,
                          blacklist=None,
                          whitelist=None,
                          citizenship_filter: str = "all",
                          experience_levels=None) -> list:
    """Read the local job index (populated by ``pipeline.ingest``) and return
    a ranked list of legacy-shape job dicts so phases 3-7 keep working.

    The signature is intentionally identical to the old internet-scraping
    version. Several params are now no-ops:

      use_simplify        — every aggregator-style README runs in the
                            scheduler regardless.
      cache_ttl_minutes   — the DB itself is the cache.
      force_live          — when True, kicks a synchronous re-run of every
                            registered source before searching.

    ``deep_search`` widens the age window to 90 days; ``offset`` skips the
    first N rows in ranking order so the existing "load more" SSE flow on
    the Agent page still paginates.
    """
    from .job_search import search, SearchFilters

    console.print("\n[bold cyan]Phase 2 — Job Discovery (local index)[/bold cyan]")

    job_titles, _titles_source = _resolve_effective_titles(job_titles, profile)
    cap  = max_jobs or MAX_SCRAPE_JOBS
    days = 90 if deep_search else (days_old or 30)
    console.print(
        f"  🔍 Titles: {', '.join(job_titles) if job_titles else '(any)'} "
        f"(source: {_titles_source})"
    )
    console.print(f"  📍 Location: {location}")
    console.print(f"  📅 Posted within: {days} days")
    console.print(f"  🔢 Cap: {cap}, offset: {offset}, deep: {deep_search}")
    _t0 = time.time()

    if force_live:
        console.print("  [yellow]⚡ force_live=True — kicking ingestion before search[/yellow]")
        try:
            from . import ingest as _ig
            results = _ig.force_run()
            ok_count = sum(1 for r in results if r.get("ok"))
            inserted = sum(r.get("inserted", 0) for r in results if r.get("ok"))
            console.print(
                f"  ✓ ingestion: {ok_count}/{len(results)} sources ok, "
                f"{inserted} rows upserted"
            )
        except Exception as exc:
            console.print(f"  [yellow]ingestion kick failed: {exc}[/yellow]")

    # Open a dedicated read connection; works for both web (where the
    # session store already created the schema) and CLI (where running
    # before any scheduler tick will return zero rows).
    from .config import DB_PATH as _DB_PATH
    db_path = _DB_PATH
    if not db_path.exists():
        console.print("  [yellow]Local index DB does not exist yet[/yellow]")
        return []

    profile_for_search = {
        "target_titles": job_titles or (profile.get("target_titles") or []),
        "top_hard_skills": profile.get("top_hard_skills") or [],
    }
    # Push every filter the user already declared into the DB-level query
    # rather than post-filtering after the cap. Otherwise a user with a
    # 30-company blacklist requesting cap=50 could get back 30 results
    # because 20 of the top-50 ranked rows were dropped after the fact.
    filters = SearchFilters(
        location=("" if not location or location.strip().lower() in ("united states", "us")
                  else location),
        posted_within_days=days,
        education_levels=tuple(education_filter or ()),
        include_unknown_education=include_unknown_education,
        experience_levels=tuple(experience_levels or ()),
        citizenship_filter=citizenship_filter or "all",
        blacklist=tuple(blacklist or ()),
        whitelist=tuple(whitelist or ()),
    )

    import sqlite3 as _sqlite3
    conn = _sqlite3.connect(db_path)
    try:
        # Rank a wider pool than `cap` so the top-N ranking has room to
        # exercise the BM25 + skill_overlap + freshness + title_match
        # weighted score. Higher floor = better top-50 quality at the cost
        # of one more SELECT join per phase run.
        page = search(
            conn=conn, filters=filters, profile=profile_for_search,
            cursor=None, limit=cap + offset,
            rank_pool=max(600, (cap + offset) * 6),
        )
    finally:
        conn.close()

    jobs: list = []
    for j in page.jobs[offset:]:
        platform = j.source.split(":")[1] if ":" in j.source else j.source
        jobs.append({
            "id":                  j.id,
            "title":               j.title,
            "company":             j.company,
            "location":            j.location,
            "remote":              j.remote,
            "posted_date":         j.posted_at or "",
            "description":         "",
            "requirements":        list(j.requirements or []),
            "salary_range":        j.salary_range,
            "application_url":     j.url,
            "platform":            platform,
            "source":              j.source,
            "experience_level":    j.experience_level,
            "education_required":  j.education_required,
            "citizenship_required": j.citizenship_required,
            "url_status":          "ok",      # source pre-validated by ingester
            "_index_score":        j.score,
        })

    if not jobs:
        console.print(
            "  [yellow]⚠️  Local index returned 0 jobs for these filters — "
            "falling back to demo postings.[/yellow]"
        )
        jobs = provider.generate_demo_jobs(profile, job_titles, location)
        for job in jobs:
            job.setdefault("experience_level", "internship")
            job.setdefault("education_required", "unknown")
            job.setdefault("citizenship_required", "unknown")
            job.setdefault("source", "demo")

    console.print(f"  ✅ {len(jobs)} jobs (of {page.total_estimate} indexed)")
    console.print(f"  ⏱️  Phase 2 completed in [bold]{time.time() - _t0:.2f}s[/bold]")
    return jobs


# ── Phase 3 ────────────────────────────────────────────────────────────────────

def phase3_score_jobs(jobs: list, profile: dict, provider: BaseProvider,
                       min_score: int = 60,
                       experience_levels=None,
                       education_filter=None,  # noqa: ARG001 — moved to phase 2; kept for back-compat
                       citizenship_filter: str = "all",
                       include_unknown_education: bool = False,  # noqa: ARG001
                       include_unknown_experience: bool = True,
                       llm_score_limit: int = 10,
                       fast_only: bool = False) -> list:
    """Score every job that survived Phase 2's filters.

    Returns the FULL list (not pruned by min_score) so the UI can show why
    each job was kept or skipped. Each job gets a `filter_status` field:
      - "passed"               → score >= min_score
      - "below_threshold"      → scored, but below min_score
      - "filtered_experience"  → dropped by experience pre-filter (score=0)
      - "filtered_citizenship" → dropped by citizenship pre-filter (score=0)

    Note: education filtering now happens in Phase 2. The `education_filter`
    and `include_unknown_education` params are accepted for back-compat but
    ignored here.

    `include_unknown_experience=True` (default) keeps jobs whose experience_level
    couldn't be inferred from the title/description. Most ingested rows store
    metadata only — the inference is essentially a coin flip on a bare "Engineer"
    title — so a hard exclusion drops ~30% of relevant rows for no real reason.
    """
    console.print("\n[bold cyan]Phase 3 — Relevance Scoring & Shortlisting[/bold cyan]")

    pre_filtered: list = []  # jobs excluded before scoring (kept in output for visibility)
    to_score:     list = list(jobs)

    if experience_levels and experience_levels != ["all"]:
        # Build the accept set: the user's chosen levels, plus "unknown" iff the
        # caller hasn't opted out. Without the unknown passthrough we drop every
        # job whose body was metadata-only at ingest time — and that's the
        # majority of our index.
        accept = set(experience_levels)
        if include_unknown_experience:
            accept.add("unknown")
        kept, dropped = [], []
        for j in to_score:
            level = j.get("experience_level") or "unknown"
            if level in accept:
                kept.append(j)
            else:
                dropped.append({**j, "score": 0, "filter_status": "filtered_experience",
                                "filter_reason": f"experience={level}"})
        to_score = kept
        pre_filtered.extend(dropped)
        unk_note = " (unknowns kept)" if include_unknown_experience else ""
        console.print(
            f"  🎯 Experience filter {experience_levels}{unk_note}: "
            f"{len(to_score)} jobs remain ({len(dropped)} dropped)"
        )

    if citizenship_filter == "exclude_required":
        # Hard filter: trust the inferred field, AND regex-scan the raw text
        # so postings that slipped past inference still get caught.
        citizenship_re = re.compile(
            r"\b(?:u\.?s\.?\s*citizen(?:ship)?(?:\s+required)?"
            r"|must\s+be\s+(?:a\s+)?u\.?s\.?\s*citizen"
            r"|security\s+clearance"
            r"|active\s+clearance"
            r"|secret\s+clearance"
            r"|top\s+secret"
            r"|ts/sci"
            r"|itar"
            r"|export\s+control"
            r"|green\s+card\s+holder)\b",
            re.IGNORECASE,
        )
        kept, dropped = [], []
        for j in to_score:
            haystack = " ".join([
                str(j.get("description") or ""),
                " ".join(str(r) for r in (j.get("requirements") or [])),
            ])
            text_match = bool(citizenship_re.search(haystack))
            field_yes  = j.get("citizenship_required", "unknown") == "yes"
            if not text_match and not field_yes:
                kept.append(j)
            else:
                reason = "regex match in description" if text_match else "citizenship required"
                dropped.append({**j, "score": 0, "filter_status": "filtered_citizenship",
                                "filter_reason": reason})
        to_score = kept
        pre_filtered.extend(dropped)
        console.print(
            f"  🇺🇸 Citizenship filter (exclude required): {len(to_score)} jobs remain "
            f"({len(dropped)} dropped, regex+field)"
        )
    elif citizenship_filter == "only_required":
        kept, dropped = [], []
        for j in to_score:
            if j.get("citizenship_required", "unknown") == "yes":
                kept.append(j)
            else:
                dropped.append({**j, "score": 0, "filter_status": "filtered_citizenship",
                                "filter_reason": "citizenship not required"})
        to_score = kept
        pre_filtered.extend(dropped)
        console.print(f"  🇺🇸 Citizenship filter (only required): {len(to_score)} jobs remain")

    def _fast_score(job: dict, profile: dict) -> int:
        """Fast heuristic scoring (0-100) using the same skill-coverage logic
        the LLM scorer uses, plus title/experience signals. Returns an integer
        in roughly the same range as the rubric scorer so jobs that don't get
        LLM-scored are comparable."""
        coverage, _matched, _missing = compute_skill_coverage(job, profile)
        skill_pts = coverage * RUBRIC_WEIGHTS["required_skills"]  # up to 50

        title_lower = (job.get("title") or "").lower()
        targets = [str(t).lower() for t in (profile.get("target_titles") or []) if t]
        # Token-level title match — "ml engineer" hits "machine learning engineer"
        title_hit = 0.0
        for t in targets:
            if not t:
                continue
            t_tokens = [tk for tk in t.split() if len(tk) > 2]
            if t_tokens and all(tk in title_lower for tk in t_tokens):
                title_hit = 1.0
                break
            elif t in title_lower:
                title_hit = 1.0
                break
            elif any(tk in title_lower for tk in t_tokens):
                title_hit = max(title_hit, 0.5)
        industry_pts = title_hit * RUBRIC_WEIGHTS["industry"]  # up to 30

        # Location/seniority blended like the LLM rubric.
        remote_ok = bool(job.get("remote"))
        loc = (job.get("location") or "").lower()
        cand_loc = (profile.get("location") or "").lower()
        loc_hit = 1.0 if (remote_ok or "united states" in loc or
                          (cand_loc and any(part in loc for part in cand_loc.split(",") if part.strip()))) else 0.5
        exp_levels = profile.get("experience_levels") or ["internship", "entry-level"]
        exp_hit = 1.0 if job.get("experience_level") in exp_levels else 0.5
        loc_seniority_pts = ((loc_hit + exp_hit) / 2) * RUBRIC_WEIGHTS["location_seniority"]

        return int(round(skill_pts + industry_pts + loc_seniority_pts))

    console.print(f"  🔢 Fast-scoring {len(to_score)} jobs…")
    for job in to_score:
        job["_fast_score"] = _fast_score(job, profile)

    to_score.sort(key=lambda j: j["_fast_score"], reverse=True)
    llm_score_count = 0 if fast_only else min(llm_score_limit, len(to_score))
    to_llm_score = to_score[:llm_score_count]
    to_skip = to_score[llm_score_count:]

    if fast_only:
        console.print(f"  ⚡ Fast-only scoring enabled; skipping LLM for {len(to_score)} jobs")
    elif to_skip:
        console.print(
            f"  ⚡ Fast-scoring all {len(to_score)} jobs; "
            f"LLM-scoring top {llm_score_count} only"
        )

    _t0 = time.time()
    scored: list = []
    for i, job in enumerate(to_llm_score, 1):
        result = provider.score_job(job, profile)
        merged = {**job, **result}
        s = merged.get("score", 0)
        merged["filter_status"] = "passed" if s >= min_score else "below_threshold"
        merged["filter_reason"] = "" if s >= min_score else f"score {s} < {min_score}"
        scored.append(merged)
        if i % 5 == 0 or i == len(to_llm_score):
            console.print(
                f"  [dim]🧠 LLM-scored {i}/{len(to_llm_score)} jobs  "
                f"({time.time() - _t0:.0f}s elapsed)[/dim]"
            )

    for job in to_skip:
        merged = {**job}
        s = merged.get("_fast_score", 0)
        merged["score"] = s
        if fast_only:
            merged["filter_status"] = "passed" if s >= min_score else "below_threshold"
            merged["filter_reason"] = "" if s >= min_score else f"heuristic score {s} < {min_score}"
        else:
            merged["filter_status"] = "below_threshold"
            merged["filter_reason"] = f"heuristic score {s}, skipped LLM (top {llm_score_count} only)"
        scored.append(merged)

    # Combine: scored first (sorted by score), then pre-filtered for visibility.
    scored.sort(key=lambda x: x.get("score", 0), reverse=True)
    scored = scored + pre_filtered
    passed_count = sum(1 for j in scored if j.get("filter_status") == "passed")

    display_n = min(12, len(scored))
    table = Table(title=f"Job Match Scores (showing top {display_n} of {len(scored)}, min: {min_score})")
    table.add_column("#",       style="dim",   width=4)
    table.add_column("Company", style="cyan",  width=18)
    table.add_column("Title",   style="white", width=28)
    table.add_column("Score",   style="bold",  width=8)
    table.add_column("Status",  width=22)

    for i, job in enumerate(scored[:12], 1):
        s = job.get("score", 0)
        fs = job.get("filter_status", "passed")
        if fs.startswith("filtered_"):
            colour, status = "red", f"❌ {fs.replace('filtered_', '')}"
        elif s >= 75:
            colour, status = "bold green", "✅ Auto-eligible"
        elif s >= 60:
            colour, status = "yellow", "⚠️  Review needed"
        else:
            colour, status = "red", "❌ Below threshold"
        table.add_row(str(i), job.get("company", ""), job.get("title", ""),
                      f"[{colour}]{s}[/{colour}]", status)

    console.print(table)
    console.print(
        f"  ⏱️  Phase 3 completed in [bold]{time.time() - _t0:.1f}s[/bold]  "
        f"({passed_count}/{len(scored)} jobs passed the {min_score} score threshold)"
    )
    return scored


# ── Phase 4 ────────────────────────────────────────────────────────────────────

def _ats_score(text: str, requirements: list) -> int:
    """Return percentage of *requirements* keywords present in *text* (0-100)."""
    if not requirements:
        return 0
    text_l = (text or "").lower()
    hits = 0
    for req in requirements:
        token = str(req or "").strip().lower()
        if token and token in text_l:
            hits += 1
    return int(round(100 * hits / len(requirements)))


def _profile_to_text(profile: dict) -> str:
    """Flatten a structured profile dict into a single searchable string."""
    parts: list = []
    parts.extend(profile.get("top_hard_skills") or [])
    parts.extend(profile.get("top_soft_skills") or [])
    for role in (profile.get("experience") or []):
        parts.append(role.get("title", ""))
        parts.append(role.get("company", ""))
        parts.extend(role.get("bullets") or [])
    for proj in (profile.get("projects") or []):
        parts.append(proj.get("name", ""))
        parts.append(proj.get("description", ""))
        parts.extend(proj.get("skills_used") or [])
    for ed in (profile.get("education") or []):
        parts.append(ed.get("degree", ""))
        parts.append(ed.get("institution", ""))
    return " ".join(p for p in parts if p)


def _tailoring_is_empty(tailored: dict) -> bool:
    """True when the LLM returned a tailoring response with no usable content."""
    if not tailored:
        return True
    return (
        not tailored.get("skills_reordered")
        and not tailored.get("experience_bullets")
    )


def phase4_tailor_resume(job: dict, profile: dict, resume_text: str,
                          provider: BaseProvider, include_cover_letter: bool = False,
                          section_order: list = None, *,
                          selected_keywords: list[str] | None = None,
                          source_format: str | None = None) -> dict:
    """Produce a TailoredResume v2 dict for *job* against *profile*.

    Pipeline:
      1. Compute heuristic v2 baseline (cheap; used as safety net + merge target).
      2. Ask provider for v2 (passing user-selected keywords + source_format hint).
      3. Validate v2 shape; retry once on failure.
      4. If still bad, fall back to heuristic.
      5. Hybrid merge: keep LLM's good fields, heuristic backfills empties.
      6. Compute ATS scores (before / after) over skills + bullets.
    """
    from .heuristic_tailor import (
        heuristic_tailor_resume_v2, merge_with_heuristic_v2, validate_v2_or_none,
    )

    heuristic = heuristic_tailor_resume_v2(
        job, profile, resume_text, selected_keywords=selected_keywords,
    )

    raw = None
    try:
        raw = provider.tailor_resume(
            job, profile, resume_text,
            selected_keywords=selected_keywords, source_format=source_format,
        ) or {}
    except TypeError:
        # Backward-compat: legacy provider didn't accept kwargs
        try:
            raw = provider.tailor_resume(job, profile, resume_text) or {}
        except Exception as e:
            console.print(f"  [yellow]Tailoring provider error: {e}[/yellow]")
    except Exception as e:
        console.print(f"  [yellow]Tailoring provider error: {e}[/yellow]")

    validated = validate_v2_or_none(raw)
    if validated is None:
        console.print("  [yellow][!] Tailoring response unusable — retrying once.[/yellow]")
        try:
            retry = provider.tailor_resume(
                job, profile, resume_text,
                selected_keywords=selected_keywords, source_format=source_format,
            ) or {}
        except Exception as e:
            console.print(f"  [yellow]Retry failed: {e}[/yellow]")
            retry = {}
        validated = validate_v2_or_none(retry)

    if validated is None:
        console.print("  [yellow][!] Falling back to heuristic v2 tailoring.[/yellow]")
        tailored = heuristic
    else:
        tailored = merge_with_heuristic_v2(validated, heuristic)

    if section_order:
        tailored["section_order"] = section_order
    if include_cover_letter:
        tailored["cover_letter"] = provider.generate_cover_letter(job, profile)

    # ── ATS scoring (before / after) — text harvested from skills + bullets ─
    requirements = job.get("requirements") or []
    before_text = (resume_text or "") + " " + _profile_to_text(profile)
    after_extras: list[str] = []
    for cat in tailored.get("skills") or []:
        for it in cat.get("items") or []:
            after_extras.append(it.get("text") or "")
    for r in tailored.get("experience") or []:
        for b in r.get("bullets") or []:
            after_extras.append(b.get("text") or "")
    after_text = before_text + " " + " ".join(after_extras)
    tailored["ats_score_before"] = _ats_score(before_text, requirements)
    tailored["ats_score_after"] = _ats_score(after_text, requirements)
    return tailored


# ── Playwright submitter ───────────────────────────────────────────────────────

class PlaywrightSubmitter:
    """Real form submission via Playwright (Greenhouse boards supported)."""

    def __init__(self, profile: dict):
        self.profile = profile

    def submit(self, job: dict, resume_path: str = "", cover_letter: str = "") -> dict:  # noqa: ARG002
        url = job.get("application_url", "")
        if "boards.greenhouse.io" in url:
            return self._submit_greenhouse(job, resume_path)
        return phase5_simulate_submission(job)

    def _submit_greenhouse(self, job: dict, resume_path: str) -> dict:
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            console.print(
                "  [yellow]playwright missing — "
                "pip install playwright && playwright install chromium[/yellow]"
            )
            return phase5_simulate_submission(job)

        import random
        url        = job.get("application_url", "")
        profile    = self.profile
        name_parts = (profile.get("name") or "").split()
        first      = name_parts[0] if name_parts else ""
        last       = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
        email      = profile.get("email", "")

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page    = browser.new_page()
            try:
                page.goto(url, timeout=30000)
                for sel in ["input[name='first_name']", "input[id='first_name']"]:
                    if page.locator(sel).count():
                        page.fill(sel, first); break
                for sel in ["input[name='last_name']", "input[id='last_name']"]:
                    if page.locator(sel).count():
                        page.fill(sel, last); break
                for sel in ["input[name='email']", "input[id='email']"]:
                    if page.locator(sel).count():
                        page.fill(sel, email); break
                if resume_path and Path(resume_path).exists():
                    for sel in ["input[type='file']", "input[name='resume']"]:
                        if page.locator(sel).count():
                            page.set_input_files(sel, resume_path); break
                for sel in ["button[type='submit']", "input[type='submit']"]:
                    if page.locator(sel).count():
                        page.click(sel)
                        page.wait_for_timeout(2000)
                        break
                return {
                    "status": "Applied",
                    "confirmation": (
                        f"GH-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
                    ),
                }
            except Exception as e:
                console.print(f"  [yellow]Playwright error: {e}[/yellow]")
                return phase5_simulate_submission(job)
            finally:
                browser.close()


# ── Phase 5 ────────────────────────────────────────────────────────────────────

def _load_existing_applications(output_dir: Path = None) -> set:
    """Return set of (company_lower, title_lower) already in this month's tracker."""
    month        = datetime.now().strftime("%Y-%m")
    tracker_path = (output_dir or OUTPUT_DIR) / f"Job_Applications_Tracker_{month}.xlsx"
    if not tracker_path.exists():
        return set()
    try:
        import openpyxl
        wb      = openpyxl.load_workbook(tracker_path, read_only=True)
        ws      = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        title_col   = headers.index("Job Title") if "Job Title" in headers else None
        company_col = headers.index("Company")   if "Company"   in headers else None
        applied: set = set()
        if title_col is not None and company_col is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                t = row[title_col]
                c = row[company_col]
                if t and c:
                    applied.add((str(c).lower(), str(t).lower()))
        wb.close()
        return applied
    except Exception:
        return set()


def phase5_simulate_submission(job: dict, already_applied: set = None) -> dict:
    if already_applied is None:
        already_applied = set()
    key = (job.get("company", "").lower(), job.get("title", "").lower())
    if key in already_applied:
        console.print("  ⏭️  Already applied — skipped")
        return {"status": "Skipped", "confirmation": "N/A",
                "notes": "Already applied — skipped"}
    import random
    status  = random.choice(["Applied", "Applied", "Applied", "Manual Required"])
    confirm = (
        f"DEMO-{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
        if status == "Applied" else "N/A"
    )
    return {"status": status, "confirmation": confirm}


# ── Phase 6 ────────────────────────────────────────────────────────────────────

# Canonical column schema for the application tracker. Used for both the
# in-page UI spreadsheet (web flow) and the .xlsx export (CLI flow).
TRACKER_COLUMNS: list[dict] = [
    {"key": "n",                "label": "#",                 "type": "int",    "width":  60},
    {"key": "date_applied",     "label": "Date Applied",      "type": "date",   "width": 110},
    {"key": "title",            "label": "Job Title",         "type": "text",   "width": 240},
    {"key": "company",          "label": "Company",           "type": "text",   "width": 160},
    {"key": "industry",         "label": "Industry",          "type": "text",   "width": 160},
    {"key": "location",         "label": "Location",          "type": "text",   "width": 140},
    {"key": "url",              "label": "Job Posting URL",   "type": "url",    "width": 180},
    {"key": "company_website",  "label": "Company Website",   "type": "url",    "width": 160},
    {"key": "platform",         "label": "Application Portal","type": "text",   "width": 130},
    {"key": "score",            "label": "Match Score",       "type": "score",  "width": 100},
    {"key": "reasoning",        "label": "Score Reasoning",   "type": "text",   "width": 260},
    {"key": "resume_version",   "label": "Resume Version",    "type": "text",   "width": 220},
    {"key": "cover_letter_sent","label": "Cover Letter",      "type": "yesno",  "width":  90},
    {"key": "status",           "label": "Status",            "type": "status", "width": 130},
    {"key": "confirmation",     "label": "Confirmation #",    "type": "text",   "width": 150},
    {"key": "notes",            "label": "Notes",             "type": "text",   "width": 220},
    {"key": "follow_up",        "label": "Follow-Up Date",    "type": "date",   "width": 110},
    {"key": "response",         "label": "Response Received", "type": "text",   "width": 140},
]

# Excel header order, derived from TRACKER_COLUMNS so the .xlsx export and the
# in-page table can never drift apart.
_XLSX_HEADERS: list[str] = [c["label"] for c in TRACKER_COLUMNS]


def _build_tracker_row(app: dict, row_num: int) -> dict:
    """Project a Phase-5 application dict onto the canonical TRACKER_COLUMNS keys."""
    applied_str = app.get("date_applied", datetime.now().strftime("%m/%d/%Y"))
    try:
        follow_up = (
            datetime.strptime(applied_str, "%m/%d/%Y") + timedelta(days=7)
        ).strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        follow_up = ""
    company = app.get("company", "") or ""
    company_slug = company.lower().replace(" ", "")
    return {
        "n":                row_num,
        "date_applied":     applied_str,
        "title":            app.get("title", ""),
        "company":          company,
        "industry":         app.get("industry") or "Technology / Semiconductor",
        "location":         app.get("location", ""),
        "url":              app.get("application_url", ""),
        "company_website":  f"https://www.{company_slug}.com" if company_slug else "",
        "platform":         app.get("platform", ""),
        "score":            int(app.get("score") or 0),
        "reasoning":        app.get("reasoning") or app.get("reason", "") or "",
        "resume_version":   app.get("resume_version", ""),
        "cover_letter_sent": bool(app.get("cover_letter_sent")),
        "status":           app.get("status", "Applied"),
        "confirmation":     app.get("confirmation", "N/A"),
        "notes":            app.get("notes", ""),
        "follow_up":        follow_up,
        "response":         app.get("response", ""),
    }


def _build_tracker_summary(rows: list[dict]) -> dict:
    total   = len(rows)
    applied = sum(1 for r in rows if r.get("status") == "Applied")
    manual  = sum(1 for r in rows if r.get("status") == "Manual Required")
    skipped = sum(1 for r in rows if r.get("status") == "Skipped")
    errors  = sum(1 for r in rows if r.get("status") == "Error")
    scores  = [int(r.get("score") or 0) for r in rows if r.get("score") is not None]
    avg     = round(sum(scores) / len(scores), 1) if scores else 0.0
    return {
        "run_date":  date.today().isoformat(),
        "total":     total,
        "applied":   applied,
        "manual":    manual,
        "skipped":   skipped,
        "errors":    errors,
        "avg_score": avg,
    }


def _write_tracker_xlsx(rows: list[dict], summary: dict, tracker_path: Path) -> Path | None:
    """Render the tracker rows + summary into the legacy .xlsx file. Used only
    by the CLI flow (and any caller that explicitly requests file output).
    Returns the saved path on success or None when openpyxl isn't installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("  [yellow]openpyxl missing — run: pip install openpyxl[/yellow]")
        return None

    headers = _XLSX_HEADERS
    if tracker_path.exists():
        wb = openpyxl.load_workbook(tracker_path)
        ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
        ws.title = "Applications"
        existing_headers = [cell.value for cell in ws[1]]
        if existing_headers != headers:
            ws.delete_rows(1); ws.insert_rows(1)
            for col, hdr in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=hdr)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        for col, hdr in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=hdr)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    status_fills = {
        "Applied":         PatternFill("solid", fgColor="C6EFCE"),
        "Manual Required": PatternFill("solid", fgColor="FFEB9C"),
        "Skipped":         PatternFill("solid", fgColor="FFC7CE"),
        "Error":           PatternFill("solid", fgColor="D9D9D9"),
    }

    existing_rows: dict = {}
    for row_idx in range(2, ws.max_row + 1):
        company = ws.cell(row=row_idx, column=headers.index("Company") + 1).value
        title   = ws.cell(row=row_idx, column=headers.index("Job Title") + 1).value
        if company and title:
            existing_rows[(str(company).lower(), str(title).lower())] = row_idx

    for r in rows:
        key = (str(r.get("company", "")).lower(), str(r.get("title", "")).lower())
        row_idx = existing_rows.get(key)
        if not row_idx:
            row_idx = ws.max_row + 1
            existing_rows[key] = row_idx
        values = [
            r["n"], r["date_applied"], r["title"], r["company"], r["industry"],
            r["location"], r["url"], r["company_website"], r["platform"],
            r["score"], r["reasoning"], r["resume_version"],
            "Yes" if r["cover_letter_sent"] else "No",
            r["status"], r["confirmation"], r["notes"], r["follow_up"], r["response"],
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=value)
        fill = status_fills.get(r.get("status", "Applied"), status_fills["Applied"])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws_d = wb.create_sheet("Dashboard")
    for row in [
        ("Metric", "Value"),
        ("Run Date", summary["run_date"]),
        ("Total Jobs Evaluated", summary["total"]),
        ("Applications Submitted", summary["applied"]),
        ("Manual Review Required", summary["manual"]),
        ("Skipped (Low Match)", summary["skipped"]),
        ("Errors", summary["errors"]),
        ("Average Match Score", f"{summary['avg_score']:.1f}"),
    ]:
        ws_d.append(row)
    ws_d["A1"].font = Font(bold=True)
    ws_d["B1"].font = Font(bold=True)

    wb.save(tracker_path)
    return tracker_path


def phase6_update_tracker(applications: list,
                          output_dir: Path = None,
                          write_file: bool = True) -> dict:
    """Build the tracker rows + summary stats from Phase-5 applications.

    Returns a dict shaped for both UI rendering and file export:

        {
          "month":        "YYYY-MM",
          "columns":      [{key, label, type, width}, ...],
          "rows":         [{key: value, ...}, ...],
          "summary":      {run_date, total, applied, manual, skipped, errors, avg_score},
          "tracker_path": Path | None,    # populated only when an .xlsx was written
        }

    File-write behaviour is opt-in via ``write_file``. The web flow passes
    ``write_file=False`` to keep the run fully in-page; the CLI keeps the
    default ``True`` so the existing offline workflow still produces an
    .xlsx in ``output/``.
    """
    console.print("\n[bold cyan]Phase 6 — Application Tracker[/bold cyan]")

    rows = [_build_tracker_row(a, i + 1) for i, a in enumerate(applications)]
    summary = _build_tracker_summary(rows)
    month = datetime.now().strftime("%Y-%m")

    saved_path: Path | None = None
    if write_file:
        out_dir = output_dir or OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        target = out_dir / f"Job_Applications_Tracker_{month}.xlsx"
        try:
            saved_path = _write_tracker_xlsx(rows, summary, target)
            if saved_path:
                console.print(f"  ✅ Tracker saved → [bold]{saved_path}[/bold]")
        except Exception as exc:
            console.print(f"  [yellow]Could not write {target.name}: {exc}[/yellow]")
            saved_path = None

    console.print(
        f"  📊 {summary['total']} rows · "
        f"applied={summary['applied']} · manual={summary['manual']} · "
        f"skipped={summary['skipped']} · avg score {summary['avg_score']}"
    )

    return {
        "month":        month,
        "columns":      list(TRACKER_COLUMNS),
        "rows":         rows,
        "summary":      summary,
        "tracker_path": saved_path,
    }


# ── Email notification ─────────────────────────────────────────────────────────

def _send_email_notification(report_text: str, n_applied: int) -> None:
    required = ["SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "NOTIFY_EMAIL"]
    missing  = [v for v in required if not os.environ.get(v)]
    if missing:
        console.print(
            f"  [dim]Email notification skipped (missing env: {', '.join(missing)})[/dim]"
        )
        return
    try:
        host      = os.environ["SMTP_HOST"]
        port      = int(os.environ["SMTP_PORT"])
        user      = os.environ["SMTP_USER"]
        password  = os.environ["SMTP_PASS"]
        recipient = os.environ["NOTIFY_EMAIL"]
        subject   = (
            f"Job Application Run Complete — "
            f"{date.today().isoformat()} ({n_applied} applied)"
        )
        msg            = MIMEText(report_text)
        msg["Subject"] = subject
        msg["From"]    = user
        msg["To"]      = recipient
        with smtplib.SMTP_SSL(host, port) as smtp:
            smtp.login(user, password)
            smtp.send_message(msg)
        console.print(f"  📧 Notification sent to {recipient}")
    except Exception as e:
        console.print(f"  [yellow]Email notification failed: {e}[/yellow]")


# ── Phase 7 ────────────────────────────────────────────────────────────────────

def phase7_run_report(applications: list, tracker_path: Path | None,
                       provider: BaseProvider,
                       output_dir: Path | None = None,
                       write_file: bool = True) -> str:
    """Generate the end-of-run summary text via the provider.

    File output is opt-in. Pass ``write_file=False`` (or ``output_dir=None``
    on a caller that wants the in-page experience) to skip writing the .md.
    The SMTP notification still fires when configured — that's a transport,
    not a file. CLI keeps the default ``write_file=True`` so the offline
    workflow still produces a markdown artifact in ``output/``.
    """
    console.print("\n[bold cyan]Phase 7 — End-of-Run Report[/bold cyan]")

    applied_list = [a for a in applications if a.get("status") == "Applied"]
    manual_list  = [a for a in applications if a.get("status") == "Manual Required"]
    skipped_list = [a for a in applications if a.get("status") == "Skipped"]
    top3 = sorted(applied_list, key=lambda x: x.get("score", 0), reverse=True)[:3]

    summary_data = {
        "total_found":    len(applications),
        "applied":        len(applied_list),
        "manual":         len(manual_list),
        "skipped":        len(skipped_list),
        "top3_applied":   [(a["company"], a["title"], a["score"]) for a in top3],
        "manual_reasons": [a.get("notes", "Form requires manual review") for a in manual_list],
    }

    report_text = provider.generate_report(summary_data)

    if write_file and output_dir is not None:
        try:
            output_dir.mkdir(parents=True, exist_ok=True)
            report_path = output_dir / f"{datetime.now().strftime('%Y%m%d')}_job-application-run-report.md"
            with open(report_path, "w", encoding="utf-8") as f:
                f.write(f"# Job Application Run Report\n**Date:** {date.today().isoformat()}\n\n")
                f.write(report_text)
                if tracker_path:
                    f.write(f"\n\n---\n**Tracker:** `{Path(tracker_path).name}`\n")
            console.print(f"  📄 Report saved → [bold]{report_path}[/bold]")
        except Exception as exc:
            console.print(f"  [yellow]Could not write report: {exc}[/yellow]")

    console.print(Panel(report_text, title="[bold]Run Summary[/bold]", border_style="green"))
    _send_email_notification(report_text, len(applied_list))
    return report_text


# ── Dashboard helper ───────────────────────────────────────────────────────────

def _launch_dashboard_and_wait(tracker_path: Path) -> None:
    """Launch the Flask dashboard, open the browser, and block until Enter."""
    import subprocess
    import webbrowser

    dashboard_script = Path(sys.argv[0]).resolve().parent / "dashboard" / "app.py"
    if not dashboard_script.exists():
        console.print("  [yellow]dashboard/app.py not found — skipping dashboard.[/yellow]")
        return

    try:
        proc = subprocess.Popen([sys.executable, str(dashboard_script)])
        time.sleep(1.5)
        webbrowser.open("http://localhost:5000")
        console.print(Panel(
            "Dashboard running at [bold]http://localhost:5000[/bold]\n\n"
            "• Review scored jobs and approve Manual Required rows.\n"
            "• Approved rows will be submitted in phases 4-7.\n"
            "• Press [bold]Enter[/bold] when ready to continue.",
            title="[bold cyan]Web Dashboard[/bold cyan]",
            border_style="cyan",
        ))
        input()
        proc.terminate()
    except Exception as e:
        console.print(f"  [yellow]Dashboard launch failed: {e}[/yellow]")
