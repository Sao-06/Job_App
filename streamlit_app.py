#!/usr/bin/env python3
"""
Streamlit Web UI for the Job Application Agent.

Run:
    python -m streamlit run streamlit_app.py

Prerequisites:
    pip install streamlit pandas
    pip install -r requirements.txt
"""

import streamlit as st
import sys
import os
import time
import threading
import traceback
import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd

# ── Page config (must be first Streamlit call) ─────────────────────────────────
st.set_page_config(
    page_title="Job Application Agent",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Root path + agent import ───────────────────────────────────────────────────
ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

try:
    import pipeline as _ag
    _ag_err = None
except Exception as _e:
    _ag = None
    _ag_err = f"{type(_e).__name__}: {_e}"

# ── Session state defaults ─────────────────────────────────────────────────────
_DEFAULTS: dict = {
    # Config
    "mode":              "ollama",
    "api_key":           "",
    "ollama_model":      "",
    "resume_text":       "",
    "latex_source":      None,
    "job_titles":        "Engineer",
    "location":          "United States",
    "threshold":         75,
    "max_apps":          10,
    "max_scrape_jobs":   50,
    "cover_letter":      False,
    "blacklist":         "",
    "whitelist":         "NVIDIA, Apple, Microsoft, Intel, IBM, Micron, Samsung, TSMC",
    # Filters
    "experience_levels": ["internship", "entry-level"],
    "education_filter":  ["bachelors"],
    "include_unknown_education": False,
    "citizenship_filter": "exclude_required",
    "use_simplify":      True,
    "days_old":          30,
    # Phase results
    "profile":      None,
    "jobs":         None,
    "scored_jobs":  None,
    "tailored_map": {},
    "applications": [],
    "report":       None,
    "tracker_path": None,
    # Pipeline state
    "phase_done":   set(),
    "phase_error":  {},
    "phase_times":  {},   # phase_n -> elapsed seconds
    "run_all":      False,
}

for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── Provider factory ───────────────────────────────────────────────────────────
def _make_provider():
    if _ag is None:
        raise RuntimeError(f"agent.py not loaded: {_ag_err}")
    mode = st.session_state.mode
    if mode == "demo":
        return _ag.DemoProvider()
    if mode == "anthropic":
        key = st.session_state.api_key.strip()
        if key:
            os.environ["ANTHROPIC_API_KEY"] = key
        return _ag.AnthropicProvider()
    if mode == "ollama":
        model = st.session_state.ollama_model.strip() or "llama3.2"
        return _ag.OllamaProvider(model=model)
    raise ValueError(f"Unknown mode: {mode}")


# ── Phase helpers ──────────────────────────────────────────────────────────────
PHASE_LABELS = {
    1: "Resume Ingestion & Profile Extraction",
    2: "Job Discovery",
    3: "Relevance Scoring & Shortlisting",
    4: "Resume Tailoring",
    5: "Application Submission",
    6: "Excel Tracker Update",
    7: "Run Report",
}


def _done(n: int) -> bool:
    return n in st.session_state.phase_done


def _errored(n: int) -> bool:
    return n in st.session_state.phase_error


def _icon(n: int) -> str:
    if _done(n):    return "✅"
    if _errored(n): return "❌"
    return "⬜"


def _run_phase(n: int, fn, *args):
    """Call fn(*args), update phase_done/phase_error, return (result, error_str)."""
    try:
        result = fn(*args)
        st.session_state.phase_error.pop(n, None)
        st.session_state.phase_done.add(n)
        return result, None
    except (Exception, SystemExit) as exc:
        st.session_state.phase_error[n] = traceback.format_exc()
        return None, str(exc)


# Loading messages per phase
_PHASE_MESSAGES = {
    1: [
        "Parsing resume — extracting skills and experience…",
        "LLM is reading your resume, hang tight…",
        "Identifying education, skills, and target roles…",
        "Still extracting profile data…",
        "Almost done parsing your resume…",
    ],
    2: [
        "Scraping LinkedIn, Indeed, Glassdoor, ZipRecruiter…",
        "Fetching job listings — network scrapes can be slow…",
        "Collecting and deduplicating postings…",
        "Merging SimplifyJobs listings…",
        "Almost done discovering jobs…",
    ],
    3: [
        "Scoring jobs against your profile…",
        "LLM is evaluating each job posting…",
        "Matching skills, titles, and experience…",
        "Running relevance scoring — this scales with job count…",
        "Almost done scoring all jobs…",
    ],
}

_GENERIC_MESSAGES = [
    "Working — please wait…",
    "Still processing…",
    "LLM is thinking…",
    "Running in background…",
    "Almost there…",
]


def _run_phase_animated(n: int, fn, *args, interval: int = 20):
    """Run fn(*args) in a background thread while cycling status messages in the UI.

    Updates session_state.phase_done / phase_error / phase_times.
    Returns (result, elapsed_seconds, error_str).
    """
    msgs = _PHASE_MESSAGES.get(n, _GENERIC_MESSAGES)

    result_holder: list = [None]
    error_holder: list = [None]   # (exception, traceback_str) or None
    done_event = threading.Event()

    def _worker():
        try:
            result_holder[0] = fn(*args)
        except Exception as exc:
            error_holder[0] = (exc, traceback.format_exc())
        finally:
            done_event.set()

    t = threading.Thread(target=_worker, daemon=True)
    t.start()

    status_box = st.empty()
    start = time.time()
    # Show first message immediately
    status_box.info(f"⏳ {msgs[0]}")
    idx = 1

    # Poll every `interval` seconds, updating the status message
    while not done_event.wait(timeout=interval):
        elapsed = int(time.time() - start)
        msg = msgs[idx % len(msgs)]
        status_box.info(f"⏳ {msg}  *(~{elapsed}s elapsed)*")
        idx += 1

    t.join()
    elapsed = round(time.time() - start, 1)
    status_box.empty()

    if error_holder[0] is not None:
        exc, tb = error_holder[0]
        st.session_state.phase_error[n] = tb
        st.session_state.phase_times[n] = elapsed
        return None, elapsed, str(exc)

    st.session_state.phase_error.pop(n, None)
    st.session_state.phase_done.add(n)
    st.session_state.phase_times[n] = elapsed
    return result_holder[0], elapsed, None


def _try_provider(phase_n: int):
    """Create the provider and return (provider, ok).
    On any error (including Ollama not running / sys.exit) stores the traceback
    in phase_error and returns (None, False) so the phase block can st.rerun().
    """
    try:
        return _make_provider(), True
    except (Exception, SystemExit) as exc:
        st.session_state.phase_error[phase_n] = traceback.format_exc()
        return None, False


# ── Visual helpers ─────────────────────────────────────────────────────────────
def _score_dot(score: int) -> str:
    thr = st.session_state.threshold
    if score >= thr: return "🟢"
    if score >= 60:  return "🟡"
    return "🔴"


def _status_dot(status: str) -> str:
    return {
        "Applied":         "🟢",
        "Manual Required": "🟡",
        "Approved":        "🔵",
        "Skipped":         "🔴",
    }.get(status, "⬜")


def _edu_icon(level: str) -> str:
    return {
        "phd":        "🎓 PhD",
        "masters":    "🎓 MS",
        "bachelors":  "🎓 BS",
        "associates": "🎓 AS",
        "high_school": "🎓 HS",
    }.get(level, "❓")


def _cit_icon(status: str) -> str:
    return {
        "yes": "🇺🇸 Required",
        "no":  "🌍 Open",
    }.get(status, "❓ Unknown")


def _loc_prefix(location: str) -> str:
    """Prefix multi-location strings with 📍."""
    if location and "," in location:
        return "📍 " + location
    return location


def _reset_pipeline():
    for k in ["profile", "jobs", "scored_jobs", "report", "tracker_path"]:
        st.session_state[k] = None
    st.session_state.tailored_map = {}
    st.session_state.applications = []
    st.session_state.phase_done   = set()
    st.session_state.phase_error  = {}
    st.session_state.phase_times  = {}
    st.session_state.run_all      = False


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Configuration")

    # ── LLM backend ───────────────────────────────────────────────────────────
    st.markdown("### LLM Backend")
    st.selectbox(
        "Mode",
        options=["ollama", "anthropic", "demo"],
        format_func=lambda x: {
            "demo":      "Demo (no API key)",
            "anthropic": "Anthropic Claude",
            "ollama":    "Ollama (local LLM)",
        }[x],
        key="mode",
    )
    if st.session_state.mode == "anthropic":
        st.text_input("ANTHROPIC_API_KEY", type="password", key="api_key",
                      placeholder="sk-ant-...")
    elif st.session_state.mode == "ollama":
        ollama_running   = False
        available_models = []
        model_data       = []
        try:
            import urllib.request as _ur
            import json as _json
            resp = _ur.urlopen("http://localhost:11434/api/tags", timeout=3)
            data = _json.loads(resp.read())
            model_data       = data.get("models", [])
            available_models = [m["name"] for m in model_data]
            ollama_running   = True
        except Exception:
            pass

        if not ollama_running:
            st.error(
                "Ollama not reachable at localhost:11434\n\n"
                "Start it with:\n```\nollama serve\n```",
                icon="🔴",
            )
            st.info("No models yet? Run:\n```\nollama pull llama3.2\n```")
        elif not available_models:
            st.warning(
                "Ollama is running but no models are pulled.\n\n"
                "```\nollama pull llama3.2\n```",
                icon="⚠️",
            )
        else:
            st.success(
                f"Ollama running — {len(available_models)} model(s) available",
                icon="🟢",
            )

            # Reset to first model if current selection is invalid
            if st.session_state.ollama_model not in available_models:
                st.session_state.ollama_model = available_models[0]

            st.selectbox(
                "Select model (smaller models are faster, larger models may improve quality)",
                options=available_models,
                index=available_models.index(st.session_state.ollama_model),
                key="ollama_model",
                help="Models currently pulled on your local Ollama installation.",
            )

            # Show model metadata beneath the selectbox
            model_info = next(
                (m for m in model_data
                 if m["name"] == st.session_state.ollama_model),
                None,
            )
            if model_info:
                size_gb = model_info.get("size", 0) / 1e9
                family  = model_info.get("details", {}).get("family", "")
                params  = model_info.get("details", {}).get("parameter_size", "")
                st.caption(
                    f"📦 {params}  |  🏷️ {family}  |  💾 {size_gb:.1f} GB"
                )

            # Warn if selected model is known to be slow
            FAST_MODELS = ["llama3.2", "llama3.1", "mistral", "phi3", "gemma3", "qwen2"]
            selected = st.session_state.ollama_model
            is_fast  = any(f in selected.lower() for f in FAST_MODELS)
            if not is_fast:
                st.warning(
                    f"'{selected}' may be slow for pipeline use. "
                    "llama3.2, mistral, or phi3 are faster.",
                    icon="🐢",
                )

            if st.button("🔄 Refresh model list",
                         use_container_width=True, key="refresh_models"):
                st.rerun()

            st.caption("To add a model: `ollama pull mistral` → Refresh.")

    # ── Resume ────────────────────────────────────────────────────────────────
    st.divider()
    st.markdown("### Resume")
    resume_src = st.radio(
        "Resume source",
        ["Demo profile", "Upload file", "Paste text"],
        label_visibility="collapsed",
    )

    if resume_src == "Upload file":
        uploaded = st.file_uploader("Upload PDF / DOCX / TXT / TEX",
                                    type=["pdf", "docx", "txt", "tex"])
        if uploaded and _ag:
            try:
                suffix = Path(uploaded.name).suffix.lower()
                raw_bytes = uploaded.read()
                if suffix == ".tex":
                    # LaTeX source — store original and convert to plaintext
                    raw_str = raw_bytes.decode("utf-8", errors="replace")
                    st.session_state.latex_source = raw_str
                    st.session_state.resume_text = _ag.latex_to_plaintext(raw_str)
                    st.success(f"LaTeX resume loaded: {uploaded.name}  "
                               f"({len(st.session_state.resume_text):,} chars plain text)")
                else:
                    # Write to a temp file so _read_resume (and pdfplumber) can open it by path
                    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                        tmp.write(raw_bytes)
                        tmp_path = Path(tmp.name)
                    text, latex_src = _ag._read_resume(tmp_path)
                    if latex_src:
                        st.session_state.latex_source = latex_src
                    # Detect silent fallback to demo resume (means extraction failed)
                    demo = _ag._build_demo_resume()
                    if text.strip() and text.strip() != demo.strip():
                        st.session_state.resume_text = text
                        latex_note = " (LaTeX detected)" if latex_src else ""
                        st.success(f"Loaded: {uploaded.name}  ({len(text):,} chars){latex_note}")
                    else:
                        st.error(
                            f"Could not extract text from **{uploaded.name}**. "
                            "Possible causes: scanned/image-only PDF, corrupted file, "
                            "or pdfplumber not installed (`pip install pdfplumber`). "
                            "Try a text-based PDF, DOCX, TXT, or TEX file instead."
                        )
            except Exception as exc:
                st.error(f"Failed to read **{uploaded.name}**: {exc}")

    elif resume_src == "Paste text":
        pasted = st.text_area("Resume text", height=150,
                              placeholder="Paste your resume here… (LaTeX source also accepted)")
        if pasted and _ag:
            if _ag.detect_latex(pasted):
                st.session_state.latex_source = pasted
                st.session_state.resume_text = _ag.latex_to_plaintext(pasted)
                st.info("LaTeX source detected — converted to plain text for parsing.")
            else:
                st.session_state.latex_source = None
                st.session_state.resume_text = pasted

    else:  # Demo profile
        if _ag and not st.session_state.resume_text:
            st.session_state.resume_text = _ag._build_demo_resume()
        if st.button("Reload demo profile", use_container_width=True):
            if _ag:
                st.session_state.resume_text = _ag._build_demo_resume()

    if st.session_state.resume_text:
        if st.session_state.latex_source:
            st.caption("LaTeX source loaded — PDF output enabled if pdflatex is installed.")
        with st.expander("Preview resume"):
            preview = st.session_state.resume_text
            st.code(
                preview[:800] + ("…" if len(preview) > 800 else ""),
                language=None,
            )

    # ── Search settings ───────────────────────────────────────────────────────
    st.divider()
    st.markdown("### Search Settings")
    st.text_input("Job titles (comma-separated)", key="job_titles")
    st.text_input("Location / Region", key="location")
    st.slider("Auto-apply threshold (score)", 50, 100, key="threshold")
    st.number_input("Max applications per run", min_value=1, max_value=50,
                    key="max_apps")
    st.number_input(
        "Max jobs to scrape (Phase 2)",
        min_value=5, max_value=200, step=5,
        key="max_scrape_jobs",
        help="Caps the total jobs collected in Phase 2. Fewer jobs = faster Phase 3 scoring.",
    )
    st.checkbox("Generate cover letters", key="cover_letter")

    st.multiselect(
        "Experience level",
        options=["internship", "entry-level", "mid-level", "senior", "unknown"],
        default=st.session_state.experience_levels,
        key="experience_levels",
    )

    st.multiselect(
        "Education requirement",
        options=["high_school", "associates", "bachelors", "masters", "phd", "unknown"],
        default=st.session_state.education_filter,
        format_func=lambda x: {
            "high_school": "High School / GED",
            "associates":  "Associate's Degree",
            "bachelors":   "Bachelor's (BS/BE)",
            "masters":     "Master's (MS/MEng)",
            "phd":         "PhD / Doctorate",
            "unknown":     "Not specified",
        }[x],
        key="education_filter",
    )

    st.checkbox(
        "Include jobs with unspecified education",
        value=st.session_state.include_unknown_education,
        key="include_unknown_education",
        help="When off, jobs whose education requirement could not be inferred are dropped. "
             "Turn on if the filter is too aggressive.",
    )

    st.checkbox(
        "Include SimplifyJobs/GitHub listings",
        value=st.session_state.use_simplify,
        key="use_simplify",
        help="Scrapes real-time internship listings from github.com/SimplifyJobs/Summer2026-Internships",
    )

    st.number_input(
        "Only show jobs posted within (days)",
        min_value=1,
        max_value=180,
        value=st.session_state.days_old,
        step=1,
        key="days_old",
        help="Default 30 days. Newest postings are ranked first.",
    )

    with st.expander("Advanced filters"):
        st.text_input("Blacklist companies (comma-sep)", key="blacklist",
                      placeholder="e.g. Acme, Globex")
        st.text_input("Priority companies (comma-sep)", key="whitelist")
        st.radio(
            "US Citizenship requirement",
            options=["all", "exclude_required", "only_required"],
            format_func=lambda x: {
                "all":              "Show all jobs",
                "exclude_required": "Exclude citizenship-required roles",
                "only_required":    "Only citizenship-required roles",
            }[x],
            key="citizenship_filter",
            index=0,
        )

    # ── Reset ─────────────────────────────────────────────────────────────────
    st.divider()
    if st.button("🔄 Reset Pipeline", use_container_width=True, type="secondary"):
        _reset_pipeline()
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN AREA
# ══════════════════════════════════════════════════════════════════════════════
st.title("💼 Job Application Agent")
st.caption("Autonomous 7-Phase Job Search & Application Pipeline")

if _ag_err:
    st.error(f"Failed to import agent.py — {_ag_err}")
    st.info("Make sure all dependencies are installed: `pip install -r requirements.txt`")
    st.stop()

tab_pipeline, tab_tracker, tab_files = st.tabs(["🚀 Pipeline", "📊 Tracker", "📁 Output Files"])


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE TAB
# ══════════════════════════════════════════════════════════════════════════════
with tab_pipeline:
    # ── Stop run_all BEFORE evaluating any should_runX condition ──────────────
    # st.rerun() restarts the script from the top, so a bottom-of-page check
    # is never reached after an error.  Moving the check here breaks the loop.
    if st.session_state.run_all and st.session_state.phase_error:
        st.session_state.run_all = False

    # Progress bar
    n_done = len(st.session_state.phase_done)
    st.progress(n_done / 7, text=f"Pipeline: {n_done} / 7 phases complete")

    col_btn, col_hint = st.columns([1, 5])
    if col_btn.button("▶ Run Full Pipeline", type="primary", use_container_width=True):
        _reset_pipeline()
        st.session_state.run_all = True
        st.rerun()
    col_hint.caption(
        "Runs all 7 phases automatically, or step through them individually below."
    )
    st.divider()

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 1 — Resume Ingestion
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(f"{_icon(1)} Phase 1 — {PHASE_LABELS[1]}",
                     expanded=not _done(1)):

        if not st.session_state.resume_text:
            st.warning("No resume loaded — select a source in the sidebar.")
        else:
            btn_p1 = None
            if not _done(1):
                btn_p1 = st.button("▶ Extract Profile", key="btn_p1")

            should_run1 = (
                (btn_p1 or (st.session_state.run_all and not _errored(1))) and
                not _done(1) and
                bool(st.session_state.resume_text)
            )

            if should_run1:
                provider, ok = _try_provider(1)
                if ok:
                    preferred = [t.strip() for t in
                                 st.session_state.job_titles.split(",") if t.strip()]
                    profile, elapsed, err = _run_phase_animated(
                        1, _ag.phase1_ingest_resume,
                        st.session_state.resume_text, provider, preferred,
                    )
                    if profile is not None:
                        st.session_state.profile = profile
                st.rerun()

            if st.session_state.profile:
                p = st.session_state.profile
                c1, c2, c3 = st.columns(3)
                c1.metric("Name",     p.get("name", "—"))
                c2.metric("Email",    p.get("email", "—"))
                c3.metric("Location", p.get("location", "—"))

                left, right = st.columns(2)
                with left:
                    titles = p.get("target_titles", [])
                    st.markdown("**Suggested target titles**")
                    if titles:
                        for t in titles:
                            st.write(f"• {t}")
                        if st.button(
                            "Use these titles for job search",
                            key="btn_use_titles",
                            help="Replaces your Search Settings job titles with these suggestions",
                        ):
                            st.session_state.job_titles = ", ".join(titles)
                            st.rerun()
                    else:
                        st.caption("No titles suggested — add preferences in Search Settings.")
                with right:
                    st.markdown("**Hard skills**")
                    st.write(", ".join(p.get("top_hard_skills", [])))
                    st.markdown("**Soft skills**")
                    st.write(", ".join(p.get("top_soft_skills", [])))

                if p.get("resume_gaps"):
                    st.warning("Resume gaps: " + " | ".join(p["resume_gaps"]))

                if st.session_state.phase_times.get(1):
                    st.caption(f"⏱️ Completed in {st.session_state.phase_times[1]:.1f}s")

        if _errored(1):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[1])

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 2 — Job Discovery
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(f"{_icon(2)} Phase 2 — {PHASE_LABELS[2]}",
                     expanded=_done(1) and not _done(2)):

        if not st.session_state.profile:
            st.info("Complete Phase 1 first.")
        else:
            # ── Cache status ───────────────────────────────────────────────────
            _cache_file = _ag.RESOURCES_DIR / "sample_jobs.json"
            if _cache_file.exists():
                import json as _json
                try:
                    _cached = _json.loads(_cache_file.read_text(encoding="utf-8"))
                    _first_url = (_cached[0].get("application_url", "") if _cached else "")
                    _is_demo   = any(
                        demo in _first_url
                        for demo in ["nvidia.com/careers", "intel.com/jobs",
                                     "microsoft.com/careers", "apple.com/jobs",
                                     "lumentum.com/careers", "micron.com/careers",
                                     "research.ibm.com", "samsung.com/us/careers"]
                    )
                except Exception:
                    _is_demo = False

                if _is_demo:
                    st.warning(
                        "**Cached jobs are demo/fake URLs** from a previous run.  "
                        "Click **Clear cache** to scrape real postings.",
                        icon="⚠️",
                    )
                else:
                    from datetime import datetime as _dt
                    _age = _dt.now().timestamp() - _cache_file.stat().st_mtime
                    _age_h = int(_age / 3600)
                    st.info(
                        f"Using cached jobs ({len(_cached)} postings, "
                        f"{_age_h}h old).  Click **Clear cache** to re-scrape.",
                        icon="📂",
                    )

                if st.button("🗑 Clear cache & re-scrape", key="btn_clear_cache"):
                    _cache_file.unlink(missing_ok=True)
                    # Also reset phase 2 so it re-runs
                    st.session_state.phase_done.discard(2)
                    st.session_state.phase_done.discard(3)
                    st.session_state.phase_done.discard(4)
                    st.session_state.phase_done.discard(5)
                    st.session_state.phase_done.discard(6)
                    st.session_state.phase_done.discard(7)
                    st.session_state.jobs        = None
                    st.session_state.scored_jobs = None
                    st.session_state.tailored_map = {}
                    st.session_state.applications = []
                    st.session_state.report       = None
                    st.rerun()

            btn_p2 = None
            if not _done(2):
                btn_p2 = st.button("▶ Discover Jobs", key="btn_p2")
                if not _cache_file.exists():
                    st.caption("Will scrape live job boards via JobSpy (LinkedIn, Indeed, Glassdoor, ZipRecruiter).")

            should_run2 = (
                (btn_p2 or (st.session_state.run_all and _done(1) and not _errored(2))) and
                not _done(2)
            )

            if should_run2:
                provider, ok = _try_provider(2)
                if ok:
                    titles = [t.strip() for t in
                              st.session_state.job_titles.split(",") if t.strip()]
                    jobs, elapsed, err = _run_phase_animated(
                        2, _ag.phase2_discover_jobs,
                        st.session_state.profile, titles,
                        st.session_state.location, provider,
                        st.session_state.use_simplify,
                        st.session_state.max_scrape_jobs,
                        st.session_state.days_old,
                    )
                    if jobs is not None:
                        st.session_state.jobs = jobs
                st.rerun()

            if st.session_state.jobs:
                jobs = st.session_state.jobs
                _mc1, _mc2, _mc3 = st.columns(3)
                _mc1.metric("Jobs found", len(jobs))
                _mc2.metric("Duplicates merged", _ag.helpers._last_merge_count)
                if st.session_state.phase_times.get(2):
                    _mc3.metric("Time", f"{st.session_state.phase_times[2]:.0f}s")
                df = pd.DataFrame([{
                    "Company":     j.get("company", ""),
                    "Role":        j.get("title", ""),
                    "Location":    _loc_prefix(j.get("location", "")),
                    "Remote":      "Yes" if j.get("remote") else "No",
                    "Experience":  j.get("experience_level", "unknown"),
                    "Education":   _edu_icon(j.get("education_required", "unknown")),
                    "Citizenship": _cit_icon(j.get("citizenship_required", "unknown")),
                    "Salary":      j.get("salary_range", ""),
                    "Platform":    j.get("platform", ""),
                    "Source":      j.get("source", ""),
                    "Date Posted": j.get("posted_date", ""),
                    "Link":        j.get("application_url", ""),
                } for j in jobs])
                st.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Link": st.column_config.LinkColumn("Link", display_text="Apply"),
                    },
                )

        if _errored(2):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[2])

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 3 — Relevance Scoring
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(f"{_icon(3)} Phase 3 — {PHASE_LABELS[3]}",
                     expanded=_done(2) and not _done(3)):

        if not st.session_state.jobs:
            st.info("Complete Phase 2 first.")
        else:
            thr = st.session_state.threshold
            btn_p3 = None
            if not _done(3):
                btn_p3 = st.button("▶ Score Jobs", key="btn_p3")
                st.caption(f"Min filter: 60  |  Auto-apply threshold: {thr}")

            should_run3 = (
                (btn_p3 or (st.session_state.run_all and _done(2) and not _errored(3))) and
                not _done(3)
            )

            if should_run3:
                provider, ok = _try_provider(3)
                if ok:
                    scored, elapsed, err = _run_phase_animated(
                        3, _ag.phase3_score_jobs,
                        st.session_state.jobs, st.session_state.profile,
                        provider, 60,
                        st.session_state.experience_levels,
                        st.session_state.education_filter,
                        st.session_state.citizenship_filter,
                        st.session_state.include_unknown_education,
                    )
                    if scored is not None:
                        st.session_state.scored_jobs = scored
                st.rerun()

            if st.session_state.scored_jobs:
                scored = st.session_state.scored_jobs
                auto   = [j for j in scored if j.get("score", 0) >= thr]
                review = [j for j in scored if 60 <= j.get("score", 0) < thr]

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Auto-eligible",    len(auto),   help=f"Score ≥ {thr}")
                m2.metric("Review needed",    len(review), help="Score 60–74")
                m3.metric("Total shortlisted", len(scored))
                if st.session_state.phase_times.get(3):
                    m4.metric("Time", f"{st.session_state.phase_times[3]:.0f}s")

                df = pd.DataFrame([{
                    "":            _score_dot(j.get("score", 0)),
                    "Company":     j.get("company", ""),
                    "Role":        j.get("title", ""),
                    "Score":       j.get("score", 0),
                    "Status":      ("Auto-eligible" if j.get("score", 0) >= thr
                                    else "Review needed"),
                    "Experience":  j.get("experience_level", "unknown"),
                    "Education":   _edu_icon(j.get("education_required", "unknown")),
                    "Citizenship": _cit_icon(j.get("citizenship_required", "unknown")),
                    "Matching Skills": ", ".join(j.get("matching_skills", [])),
                    "Missing Skills":  ", ".join(j.get("missing_skills", [])),
                    "Salary":      j.get("salary_range", ""),
                    "Link":        j.get("application_url", ""),
                } for j in scored])
                st.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Link": st.column_config.LinkColumn("Link", display_text="Apply"),
                    },
                )

        if _errored(3):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[3])

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 4 — Resume Tailoring
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(f"{_icon(4)} Phase 4 — {PHASE_LABELS[4]}",
                     expanded=_done(3) and not _done(4)):

        if not st.session_state.scored_jobs:
            st.info("Complete Phase 3 first.")
        else:
            thr      = st.session_state.threshold
            top_jobs = [j for j in st.session_state.scored_jobs
                        if j.get("score", 0) >= thr][:st.session_state.max_apps]

            btn_p4 = None
            if not _done(4):
                btn_p4 = st.button("▶ Tailor Resumes", key="btn_p4")
                st.caption(
                    f"Tailoring for {len(top_jobs)} auto-eligible job(s) "
                    f"(score ≥ {thr}, max {st.session_state.max_apps})."
                )

            should_run4 = (
                (btn_p4 or (st.session_state.run_all and _done(3) and not _errored(4))) and
                not _done(4) and
                bool(top_jobs)
            )

            if should_run4:
                prog = st.progress(0, text="Tailoring resumes…")
                try:
                    provider, ok = _try_provider(4)
                    if not ok:
                        raise RuntimeError(st.session_state.phase_error.get(4, "Provider error"))
                    tailored_map = {}
                    total = len(top_jobs)
                    for i, job in enumerate(top_jobs, 1):
                        prog.progress(
                            i / total,
                            text=f"Tailoring ({i}/{total}): "
                                 f"{job.get('company')} — {job.get('title')}",
                        )
                        result = _ag.phase4_tailor_resume(
                            job, st.session_state.profile,
                            st.session_state.resume_text, provider,
                            include_cover_letter=st.session_state.cover_letter,
                        )
                        # Generate the resume file immediately so every job
                        # has a downloadable artifact after Phase 4 (not
                        # deferred to Phase 5+6).
                        _ag.OUTPUT_DIR.mkdir(exist_ok=True)
                        resume_file = _ag._save_tailored_resume(
                            job, result, st.session_state.profile,
                            st.session_state.latex_source,
                        )
                        tailored_map[job.get("id", job.get("title", ""))] = {
                            "job": job,
                            "tailored": result,
                            "resume_file": resume_file,
                        }
                    st.session_state.tailored_map = tailored_map
                    st.session_state.phase_done.add(4)
                    st.session_state.phase_error.pop(4, None)
                except Exception as exc:
                    st.session_state.phase_error[4] = traceback.format_exc()
                finally:
                    prog.empty()
                st.rerun()

            if st.session_state.tailored_map:
                for jk, data in st.session_state.tailored_map.items():
                    job = data["job"]
                    t   = data["tailored"]
                    with st.expander(
                        f"**{job.get('company')}** — {job.get('title')}  "
                        f"(score: {job.get('score')})"
                    ):
                        skills_raw = t.get("skills_reordered", [])
                        skills_str = [s if isinstance(s, str) else str(s) for s in skills_raw]
                        st.markdown("**Skills (ATS-ordered):** " + " | ".join(skills_str))
                        section_raw = t.get("section_order", [])
                        section_str = [s if isinstance(s, str) else str(s) for s in section_raw]
                        st.markdown("**Section order:** " + " → ".join(section_str))
                        ats_raw = t.get("ats_keywords_missing", [])
                        ats_str = [s if isinstance(s, str) else str(s) for s in ats_raw]
                        if ats_str:
                            st.warning("ATS gaps — consider adding: " + ", ".join(ats_str))
                        if t.get("cover_letter"):
                            st.markdown("---")
                            st.markdown("**Cover Letter**")
                            st.text(t["cover_letter"])

                        # ── Download buttons ───────────────────────────────
                        latex_src    = st.session_state.latex_source
                        resume_file  = data.get("resume_file")
                        resume_path  = (_ag.OUTPUT_DIR / resume_file) if (_ag and resume_file) else None

                        if latex_src and _ag:
                            safe = lambda s: __import__("re").sub(
                                r"[^a-zA-Z0-9_\-]", "_", s)
                            base = (
                                f"{safe(_ag.OWNER_NAME)}_Resume"
                                f"_{safe(job.get('company',''))}"
                                f"_{safe(job.get('title',''))}"
                            )
                            tailored_latex = _ag.apply_tailoring_to_latex(latex_src, t, job)
                            dl_col1, dl_col2 = st.columns(2)
                            dl_col1.download_button(
                                "⬇ Download .tex",
                                data=tailored_latex.encode("utf-8"),
                                file_name=base + ".tex",
                                mime="text/plain",
                                key=f"dl_tex_{jk}",
                            )
                            # Try to compile PDF on-the-fly
                            import tempfile as _tmp
                            pdf_tmp = Path(_tmp.mktemp(suffix=".pdf"))
                            compiled = _ag.compile_latex_to_pdf(tailored_latex, pdf_tmp)
                            if compiled and pdf_tmp.exists():
                                dl_col2.download_button(
                                    "⬇ Download .pdf",
                                    data=pdf_tmp.read_bytes(),
                                    file_name=base + ".pdf",
                                    mime="application/pdf",
                                    key=f"dl_pdf_{jk}",
                                )
                        elif resume_path and resume_path.exists():
                            # Plain-text path: serve the .txt file written by Phase 4.
                            mime = (
                                "application/pdf" if resume_path.suffix.lower() == ".pdf"
                                else "text/plain"
                            )
                            st.download_button(
                                f"⬇ Download {resume_path.suffix.lstrip('.').upper()}",
                                data=resume_path.read_bytes(),
                                file_name=resume_path.name,
                                mime=mime,
                                key=f"dl_resume_{jk}",
                            )
                        else:
                            st.info("Resume file not generated — check Phase 4 logs.")

        if _errored(4):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[4])

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 5 & 6 — Submit + Track
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(
        f"{_icon(5)} Phase 5 & 6 — {PHASE_LABELS[5]} & {PHASE_LABELS[6]}",
        expanded=_done(4) and not _done(5),
    ):
        if not st.session_state.tailored_map:
            st.info("Complete Phase 4 first.")
        else:
            btn_p56 = None
            if not _done(5):
                btn_p56 = st.button("▶ Submit & Track", key="btn_p56", type="primary")
                st.caption(
                    "Simulates submissions (demo) and writes the Excel tracker.  "
                    "For real Playwright submission run `python agent.py --real-apply`."
                )

            should_run56 = (
                (btn_p56 or (st.session_state.run_all and _done(4) and not _errored(5))) and
                not _done(5)
            )

            if should_run56:
                with st.spinner("Submitting applications and building tracker…"):
                    try:
                        already_applied = _ag._load_existing_applications()
                        applications    = []
                        processed_ids   = set()

                        for jk, data in st.session_state.tailored_map.items():
                            job     = data["job"]
                            tailored = data["tailored"]
                            result  = _ag.phase5_simulate_submission(job, already_applied)

                            # Reuse the file generated in Phase 4 if it exists;
                            # only regenerate when missing.
                            _ag.OUTPUT_DIR.mkdir(exist_ok=True)
                            resume_file = data.get("resume_file")
                            if not resume_file or not (_ag.OUTPUT_DIR / resume_file).exists():
                                resume_file = _ag._save_tailored_resume(
                                    job, tailored, st.session_state.profile,
                                    st.session_state.latex_source,
                                )
                            processed_ids.add(job.get("id"))

                            applications.append({
                                **job,
                                "date_applied":      datetime.now().strftime("%m/%d/%Y"),
                                "resume_version":    resume_file,
                                "cover_letter_sent": bool(tailored.get("cover_letter")),
                                "status":            result["status"],
                                "confirmation":      result["confirmation"],
                                "notes": (
                                    "ATS gaps: " +
                                    ", ".join(tailored.get("ats_keywords_missing", [])[:2])
                                    if tailored.get("ats_keywords_missing") else ""
                                ),
                            })

                        # Add skipped jobs (scored but below threshold or max reached)
                        for job in st.session_state.scored_jobs:
                            if job.get("id") not in processed_ids:
                                applications.append({
                                    **job,
                                    "date_applied":      datetime.now().strftime("%m/%d/%Y"),
                                    "resume_version":    "",
                                    "cover_letter_sent": False,
                                    "status":            "Skipped",
                                    "confirmation":      "N/A",
                                    "notes": (
                                        f"Score {job.get('score', 0)} below threshold "
                                        f"or max applications reached"
                                    ),
                                })

                        tracker_path = _ag.phase6_update_tracker(applications)
                        st.session_state.applications = applications
                        st.session_state.tracker_path = tracker_path
                        st.session_state.phase_done.add(5)
                        st.session_state.phase_done.add(6)
                        st.session_state.phase_error.pop(5, None)
                        st.session_state.phase_error.pop(6, None)

                    except Exception as exc:
                        st.session_state.phase_error[5] = traceback.format_exc()
                st.rerun()

            if st.session_state.applications:
                apps = st.session_state.applications
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Applied",         sum(1 for a in apps if a.get("status") == "Applied"))
                m2.metric("Manual Required", sum(1 for a in apps if a.get("status") == "Manual Required"))
                m3.metric("Skipped",         sum(1 for a in apps if a.get("status") == "Skipped"))
                m4.metric("Total processed", len(apps))

                df = pd.DataFrame([{
                    "":            _status_dot(a.get("status", "")),
                    "Company":     a.get("company", ""),
                    "Role":        a.get("title", ""),
                    "Score":       a.get("score", 0),
                    "Status":      a.get("status", ""),
                    "Experience":  a.get("experience_level", "unknown"),
                    "Education":   _edu_icon(a.get("education_required", "unknown")),
                    "Citizenship": _cit_icon(a.get("citizenship_required", "unknown")),
                    "Confirmation": a.get("confirmation", ""),
                    "Resume File":  a.get("resume_version", ""),
                    "Link":         a.get("application_url", ""),
                } for a in apps])
                st.dataframe(
                    df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Link": st.column_config.LinkColumn("Link", display_text="Apply"),
                    },
                )

                if st.session_state.tracker_path:
                    st.success(f"Tracker saved → `{st.session_state.tracker_path}`")

        if _errored(5):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[5])

    # ─────────────────────────────────────────────────────────────────────────
    # Phase 7 — Run Report
    # ─────────────────────────────────────────────────────────────────────────
    with st.expander(f"{_icon(7)} Phase 7 — {PHASE_LABELS[7]}",
                     expanded=_done(5) and not _done(7)):

        if not st.session_state.applications:
            st.info("Complete Phases 5 & 6 first.")
        else:
            btn_p7 = None
            if not _done(7):
                btn_p7 = st.button("▶ Generate Report", key="btn_p7")

            should_run7 = (
                (btn_p7 or (st.session_state.run_all and _done(5) and not _errored(7))) and
                not _done(7)
            )

            if should_run7:
                with st.spinner("Generating end-of-run report…"):
                    provider, ok = _try_provider(7)
                    if ok:
                        report, err = _run_phase(
                            7, _ag.phase7_run_report,
                            st.session_state.applications,
                            st.session_state.tracker_path,
                            provider,
                        )
                        if report is not None:
                            st.session_state.report = report
                    st.session_state.run_all = False  # pipeline complete
                st.rerun()

            if st.session_state.report:
                st.markdown("### Run Summary")
                st.text(st.session_state.report)

                report_md = (
                    f"# Job Application Run Report\n"
                    f"**Date:** {datetime.now().date().isoformat()}\n\n"
                    f"{st.session_state.report}"
                )
                col_d1, col_d2 = st.columns(2)
                col_d1.download_button(
                    "⬇ Download Report (.md)",
                    data=report_md,
                    file_name=(
                        f"{datetime.now().strftime('%Y%m%d')}"
                        "_job-application-run-report.md"
                    ),
                    mime="text/markdown",
                )
                if st.session_state.tracker_path:
                    with open(st.session_state.tracker_path, "rb") as fh:
                        col_d2.download_button(
                            "⬇ Download Tracker (.xlsx)",
                            data=fh.read(),
                            file_name=Path(st.session_state.tracker_path).name,
                            mime=(
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet"
                            ),
                        )

        if _errored(7):
            with st.expander("Error details"):
                st.code(st.session_state.phase_error[7])

    if st.session_state.phase_error and not st.session_state.run_all:
        st.warning("One or more phases have errors. Fix the issue and click the phase button to retry.")


# ══════════════════════════════════════════════════════════════════════════════
# TRACKER TAB
# ══════════════════════════════════════════════════════════════════════════════
with tab_tracker:
    st.header("Application Tracker")

    output_dir   = _ag.OUTPUT_DIR if _ag else ROOT / "output"
    month        = datetime.now().strftime("%Y-%m")
    tracker_path = output_dir / f"Job_Applications_Tracker_{month}.xlsx"

    col_r, col_p = st.columns([1, 5])
    if col_r.button("🔄 Refresh", use_container_width=True, key="tracker_refresh"):
        st.rerun()
    col_p.caption(f"Reading: `{tracker_path}`")

    if not tracker_path.exists():
        st.info("No tracker found for this month. Run the pipeline to create one.")
    else:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(tracker_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            rows = [
                dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True)
                if any(row)
            ]
            wb.close()

            if not rows:
                st.info("Tracker is empty.")
            else:
                # Summary metrics
                statuses = [r.get("Status", "") for r in rows]
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Applied",         statuses.count("Applied"))
                m2.metric("Manual Required", statuses.count("Manual Required"))
                m3.metric("Approved",        statuses.count("Approved"))
                m4.metric("Skipped",         statuses.count("Skipped"))

                # Full table with status dot
                df = pd.DataFrame(rows)
                if "Status" in df.columns:
                    df.insert(0, "", df["Status"].map(_status_dot).fillna("⬜"))
                st.dataframe(df, use_container_width=True, hide_index=True)

                # Approve manual-review rows
                manual = [r for r in rows if r.get("Status") == "Manual Required"]
                if manual:
                    st.subheader("Pending Manual Review")
                    for r in manual:
                        c1, c2, c3, c4 = st.columns([2, 3, 1, 1])
                        c1.write(f"**{r.get('Company', '')}**")
                        c2.write(
                            f"{r.get('Job Title', '')}  |  "
                            f"Score: {r.get('Match Score', '')}"
                        )
                        url = r.get("Job Posting URL", "")
                        if url:
                            c3.link_button("View", url)
                        row_num = r.get("#")
                        if row_num and c4.button(
                            "Approve", key=f"approve_{row_num}", type="primary"
                        ):
                            wb2 = openpyxl.load_workbook(tracker_path)
                            ws2 = wb2.active
                            hdrs = [cell.value for cell in ws2[1]]
                            sc = (hdrs.index("Status") + 1
                                  if "Status" in hdrs else None)
                            if sc:
                                target = int(row_num) + 1
                                ws2.cell(row=target, column=sc).value = "Approved"
                                from openpyxl.styles import PatternFill
                                blue = PatternFill("solid", fgColor="BDD7EE")
                                for ci in range(1, len(hdrs) + 1):
                                    ws2.cell(row=target, column=ci).fill = blue
                                wb2.save(tracker_path)
                            st.success(
                                f"Approved: {r.get('Company')} — {r.get('Job Title')}"
                            )
                            st.rerun()

                # Download button
                with open(tracker_path, "rb") as fh:
                    st.download_button(
                        "⬇ Download Tracker (.xlsx)",
                        data=fh.read(),
                        file_name=tracker_path.name,
                        mime=(
                            "application/vnd.openxmlformats-officedocument"
                            ".spreadsheetml.sheet"
                        ),
                        key="tracker_dl",
                    )

        except Exception as exc:
            st.error(f"Failed to load tracker: {exc}")
            with st.expander("Details"):
                st.code(traceback.format_exc())


# ══════════════════════════════════════════════════════════════════════════════
# OUTPUT FILES TAB
# ══════════════════════════════════════════════════════════════════════════════
with tab_files:
    st.header("Output Files")

    output_dir = _ag.OUTPUT_DIR if _ag else ROOT / "output"

    if st.button("🔄 Refresh file list", key="files_refresh"):
        st.rerun()

    if not output_dir.exists() or not any(output_dir.iterdir()):
        st.info(
            "No output files yet. Run the pipeline to generate tailored resumes, "
            "the tracker, and the run report."
        )
    else:
        files    = sorted(output_dir.iterdir(),
                          key=lambda p: p.stat().st_mtime, reverse=True)
        trackers = [f for f in files if "Tracker" in f.name]
        resumes  = [f for f in files if "Resume" in f.name]
        covers   = [f for f in files if "CoverLetter" in f.name]
        reports  = [f for f in files if "report" in f.name.lower()]
        others   = [f for f in files
                    if f not in trackers + resumes + covers + reports]

        def _file_section(title: str, file_list: list):
            if not file_list:
                return
            st.markdown(f"#### {title}")
            for fp in file_list:
                size_kb = max(fp.stat().st_size // 1024, 1)
                mtime   = datetime.fromtimestamp(
                    fp.stat().st_mtime
                ).strftime("%Y-%m-%d %H:%M")
                c1, c2 = st.columns([5, 1])
                c1.write(f"`{fp.name}`  —  {size_kb} KB  |  modified {mtime}")
                with open(fp, "rb") as fh:
                    c2.download_button(
                        "⬇",
                        data=fh.read(),
                        file_name=fp.name,
                        key=f"dl_{fp.name}",
                        use_container_width=True,
                    )

        _file_section("Trackers",        trackers)
        _file_section("Tailored Resumes", resumes)
        _file_section("Cover Letters",    covers)
        _file_section("Run Reports",      reports)
        _file_section("Other",            others)

    if st.session_state.resume_text:
        st.divider()
        st.subheader("Current Resume (loaded in sidebar)")
        st.text(st.session_state.resume_text)
