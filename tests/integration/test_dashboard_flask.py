"""Integration tests for dashboard/app.py — the Flask tracker UI."""
import importlib
import sys

import openpyxl
import pytest

pytestmark = pytest.mark.integration


def _load_dashboard(tmp_path, monkeypatch):
    """Re-import dashboard.app so any env vars set on monkeypatch are picked
    up at module import time. monkeypatch.delitem auto-restores the
    sys.modules entry on teardown so the reload doesn't leak across tests.
    """
    monkeypatch.delitem(sys.modules, "dashboard.app", raising=False)
    monkeypatch.delitem(sys.modules, "dashboard", raising=False)
    dashboard_pkg = importlib.import_module("dashboard.app")
    monkeypatch.setattr(dashboard_pkg, "OUTPUT_DIR", tmp_path)
    return dashboard_pkg


@pytest.fixture
def dashboard_client(tmp_path, monkeypatch):
    """Fresh dashboard module pointed at a tmp output directory, with no
    DASHBOARD_TOKEN gate."""
    monkeypatch.delenv("DASHBOARD_TOKEN", raising=False)
    monkeypatch.delenv("DASHBOARD_HOST", raising=False)
    yield _load_dashboard(tmp_path, monkeypatch), tmp_path


@pytest.fixture
def dashboard_client_with_token(tmp_path, monkeypatch):
    """Dashboard module wired up with DASHBOARD_TOKEN=secret-123 so the
    token-gate tests share a single import."""
    monkeypatch.setenv("DASHBOARD_TOKEN", "secret-123")
    yield _load_dashboard(tmp_path, monkeypatch), tmp_path


def _seed_tracker_xlsx(tmp_path):
    """Create a minimal monthly tracker xlsx in *tmp_path* so the dashboard
    has rows to render."""
    from datetime import datetime
    month = datetime.now().strftime("%Y-%m")
    path = tmp_path / f"Job_Applications_Tracker_{month}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = ["#", "Date Applied", "Job Title", "Company", "Industry",
               "Location", "Job Posting URL", "Company Website",
               "Application Portal", "Match Score", "Score Reasoning",
               "Resume Version", "Cover Letter Sent", "Status",
               "Confirmation #", "Notes", "Follow-Up Date", "Response Received"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    rows = [
        [1, "01/15/2026", "FPGA Intern", "Acme", "Tech", "Remote",
         "https://acme/job", "https://acme.com", "Greenhouse", 85, "",
         "resume.pdf", "No", "Manual Required", "N/A", "", "01/22/2026", ""],
        [2, "01/16/2026", "Photonics Intern", "Bravo", "Tech", "SF",
         "https://b/j", "https://b.com", "LinkedIn", 78, "",
         "", "No", "Applied", "C-1", "", "01/23/2026", ""],
    ]
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


class TestDashboardIndex:
    def test_index_renders_table(self, dashboard_client):
        dashboard_pkg, tmp_path = dashboard_client
        _seed_tracker_xlsx(tmp_path)
        with dashboard_pkg.app.test_client() as client:
            r = client.get("/")
        assert r.status_code == 200
        assert b"Acme" in r.data
        assert b"Bravo" in r.data
        assert b"Manual Review" in r.data or b"badge-manual" in r.data

    def test_index_when_no_tracker(self, dashboard_client):
        dashboard_pkg, _ = dashboard_client
        with dashboard_pkg.app.test_client() as client:
            r = client.get("/")
        assert r.status_code == 200


class TestApprovalFlow:
    def test_approve_flips_manual_to_approved(self, dashboard_client):
        dashboard_pkg, tmp_path = dashboard_client
        path = _seed_tracker_xlsx(tmp_path)
        with dashboard_pkg.app.test_client() as client:
            client.get("/")
            with client.session_transaction() as s:
                csrf = s.get("_csrf")
            r = client.post("/approve/1", data={"csrf_token": csrf})
        assert r.status_code in (200, 302)
        wb = openpyxl.load_workbook(path)
        try:
            ws = wb["Applications"]
            headers = [c.value for c in ws[1]]
            status_col = headers.index("Status") + 1
            status = ws.cell(row=2, column=status_col).value
            assert status == "Approved"
        finally:
            wb.close()

    def test_approve_rejects_invalid_csrf(self, dashboard_client):
        dashboard_pkg, tmp_path = dashboard_client
        _seed_tracker_xlsx(tmp_path)
        with dashboard_pkg.app.test_client() as client:
            client.get("/")
            r = client.post("/approve/1", data={"csrf_token": "wrong-token"})
        assert r.status_code == 400

    def test_approve_rejects_already_applied(self, dashboard_client):
        dashboard_pkg, tmp_path = dashboard_client
        _seed_tracker_xlsx(tmp_path)
        with dashboard_pkg.app.test_client() as client:
            client.get("/")
            with client.session_transaction() as s:
                csrf = s.get("_csrf")
            r = client.post("/approve/2", data={"csrf_token": csrf})
        assert r.status_code == 409


class TestDashboardTokenGate:
    def test_blocks_when_token_required(self, dashboard_client_with_token):
        dashboard_pkg, _ = dashboard_client_with_token
        with dashboard_pkg.app.test_client() as client:
            r = client.get("/")
        assert r.status_code == 401

    def test_passes_with_correct_token(self, dashboard_client_with_token):
        dashboard_pkg, _ = dashboard_client_with_token
        with dashboard_pkg.app.test_client() as client:
            r = client.get("/?token=secret-123")
        assert r.status_code == 200

    def test_passes_with_correct_header(self, dashboard_client_with_token):
        dashboard_pkg, _ = dashboard_client_with_token
        with dashboard_pkg.app.test_client() as client:
            r = client.get("/", headers={"X-Dashboard-Token": "secret-123"})
        assert r.status_code == 200
