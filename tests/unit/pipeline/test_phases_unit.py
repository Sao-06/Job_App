"""Tests for pipeline.phases — happy-path of each phase using FakeProvider.

These tests focus on the data-flow contract of each phase function:
what shape it returns, what files it writes, what filters it applies. The
LLM responses are entirely canned via FakeProvider — no real network or
LLM calls.
"""
import json
from pathlib import Path

import openpyxl
import pytest

from pipeline.phases import (
    _ats_score,
    _filter_by_posting_age,
    _load_existing_applications,
    _profile_to_text,
    _resolve_effective_titles,
    _sort_newest_first,
    _tailoring_is_empty,
    phase4_tailor_resume,
    phase5_simulate_submission,
    phase6_update_tracker,
    phase7_run_report,
)
from tests.fakes import FakeProvider

pytestmark = pytest.mark.unit


# ── _ats_score ──────────────────────────────────────────────────────────────


class TestAtsScore:
    def test_full_match(self):
        assert _ats_score("verilog python fpga", ["Verilog", "Python", "FPGA"]) == 100

    def test_no_match(self):
        assert _ats_score("photonics", ["Verilog", "Python"]) == 0

    def test_partial(self):
        assert _ats_score("verilog matlab", ["Verilog", "Python", "FPGA"]) == 33

    def test_empty_requirements_returns_zero(self):
        assert _ats_score("any text", []) == 0


# ── _profile_to_text ────────────────────────────────────────────────────────


class TestProfileToText:
    def test_aggregates_all_fields(self):
        profile = {
            "top_hard_skills": ["Python"],
            "top_soft_skills": ["Teamwork"],
            "experience": [{"title": "Intern", "company": "Acme",
                             "bullets": ["did stuff", "fixed a bug"]}],
            "projects": [{"name": "ALU", "description": "8-bit ALU",
                          "skills_used": ["Verilog"]}],
            "education": [{"degree": "B.S.", "institution": "MIT"}],
        }
        text = _profile_to_text(profile)
        for needle in ("Python", "Teamwork", "Intern", "Acme", "did stuff",
                       "ALU", "Verilog", "B.S.", "MIT"):
            assert needle in text


# ── _filter_by_posting_age ──────────────────────────────────────────────────


class TestFilterByAge:
    def test_keeps_recent(self):
        from datetime import datetime, timedelta
        recent = (datetime.now() - timedelta(days=5)).strftime("%Y-%m-%d")
        old = (datetime.now() - timedelta(days=60)).strftime("%Y-%m-%d")
        jobs = [{"posted_date": recent}, {"posted_date": old}]
        kept = _filter_by_posting_age(jobs, days_old=30)
        assert len(kept) == 1
        assert kept[0]["posted_date"] == recent

    def test_keeps_undated_jobs(self):
        jobs = [{"posted_date": ""}, {"posted_date": None}, {"foo": "bar"}]
        kept = _filter_by_posting_age(jobs, days_old=30)
        assert len(kept) == 3

    def test_zero_days_old_passthrough(self):
        jobs = [{"posted_date": "1999-01-01"}]
        kept = _filter_by_posting_age(jobs, days_old=0)
        assert kept == jobs


# ── _resolve_effective_titles ───────────────────────────────────────────────


class TestResolveEffectiveTitles:
    def test_user_titles_kept(self):
        titles, src = _resolve_effective_titles(
            ["RF Engineering Intern"],
            {"target_titles": ["FPGA Intern"]}
        )
        # User input + phase-1 → merged.
        assert "RF Engineering Intern" in titles
        assert src == "merged"

    def test_only_default_engineer_falls_to_phase1(self):
        titles, src = _resolve_effective_titles(
            ["Engineer"],
            {"target_titles": ["FPGA Engineering Intern"]}
        )
        assert src == "phase1"
        assert titles == ["FPGA Engineering Intern"]

    def test_empty_user_titles_falls_to_phase1(self):
        titles, src = _resolve_effective_titles(
            [],
            {"target_titles": ["FPGA Engineering Intern"]}
        )
        assert src == "phase1"
        assert titles == ["FPGA Engineering Intern"]

    def test_no_phase1_keeps_user(self):
        titles, src = _resolve_effective_titles(["FPGA Intern"], {})
        assert src == "user"
        assert titles == ["FPGA Intern"]


# ── _tailoring_is_empty ─────────────────────────────────────────────────────


class TestTailoringIsEmpty:
    def test_none_is_empty(self):
        assert _tailoring_is_empty(None) is True
        assert _tailoring_is_empty({}) is True

    def test_skills_only_not_empty(self):
        assert _tailoring_is_empty({"skills_reordered": ["X"]}) is False

    def test_bullets_only_not_empty(self):
        assert _tailoring_is_empty({"experience_bullets": [{"role": "x"}]}) is False


# ── phase4_tailor_resume ────────────────────────────────────────────────────


class TestPhase4Tailor:
    def test_returns_ats_scores(self):
        provider = FakeProvider(tailored={
            "skills_reordered": ["Verilog", "Python"],
            "experience_bullets": [],
            "ats_keywords_missing": [],
            "section_order": ["Skills"],
        })
        job = {"title": "FPGA Intern", "company": "Acme",
               "requirements": ["Verilog", "Python"]}
        profile = {"top_hard_skills": ["Verilog", "Python"]}
        out = phase4_tailor_resume(job, profile, "Verilog Python", provider)
        assert "ats_score_before" in out
        assert "ats_score_after" in out

    def test_falls_back_when_provider_returns_empty(self):
        provider = FakeProvider(tailored={})
        job = {"title": "Eng", "company": "Co",
               "requirements": ["Verilog"]}
        profile = {"top_hard_skills": ["Python"]}
        out = phase4_tailor_resume(job, profile, "", provider)
        # Fallback synthesises skills + section_order.
        assert "skills_reordered" in out
        assert "section_order" in out

    def test_includes_cover_letter_when_requested(self):
        provider = FakeProvider(cover_letter="Dear Acme...")
        job = {"title": "FPGA Intern", "company": "Acme",
               "requirements": ["Verilog"]}
        profile = {"top_hard_skills": ["Verilog"]}
        out = phase4_tailor_resume(job, profile, "", provider,
                                    include_cover_letter=True)
        assert out["cover_letter"] == "Dear Acme..."

    def test_section_order_override(self):
        provider = FakeProvider()
        job = {"title": "X", "company": "Y", "requirements": []}
        out = phase4_tailor_resume(job, {}, "", provider,
                                    section_order=["Education", "Skills"])
        assert out["section_order"] == ["Education", "Skills"]


# ── phase5_simulate_submission ──────────────────────────────────────────────


class TestPhase5Simulate:
    def test_returns_status(self):
        out = phase5_simulate_submission({"company": "Acme", "title": "Intern"})
        assert out["status"] in ("Applied", "Manual Required", "Skipped")

    def test_already_applied_skipped(self):
        already = {("acme", "intern")}
        out = phase5_simulate_submission(
            {"company": "Acme", "title": "Intern"}, already_applied=already
        )
        assert out["status"] == "Skipped"

    def test_load_existing_applications_empty_when_no_tracker(self, tmp_path):
        assert _load_existing_applications(tmp_path) == set()


# ── phase6_update_tracker ───────────────────────────────────────────────────


class TestPhase6Tracker:
    def test_returns_dict_shape(self, tmp_path):
        apps = [
            {"company": "Acme", "title": "FPGA Intern",
             "location": "Remote", "score": 85, "status": "Applied",
             "date_applied": "01/15/2026", "application_url": "https://acme/job",
             "platform": "Greenhouse", "resume_version": "resume.pdf",
             "cover_letter_sent": False, "confirmation": "C-1",
             "notes": "test note"},
            {"company": "BetaCo", "title": "Photonics", "location": "SF",
             "score": 72, "status": "Manual Required",
             "date_applied": "01/16/2026", "application_url": "https://b/j",
             "platform": "LinkedIn", "resume_version": "",
             "cover_letter_sent": False, "confirmation": "N/A", "notes": ""},
        ]
        out = phase6_update_tracker(apps, output_dir=tmp_path)
        assert isinstance(out, dict)
        for k in ("month", "columns", "rows", "summary", "tracker_path"):
            assert k in out
        # Two app rows in.
        assert len(out["rows"]) == 2
        companies = {r.get("company") for r in out["rows"]}
        assert {"Acme", "BetaCo"}.issubset(companies)

    def test_writes_xlsx_when_write_file_true(self, tmp_path):
        apps = [{"company": "A", "title": "T", "score": 85, "status": "Applied",
                 "date_applied": "01/15/2026", "application_url": "",
                 "platform": "", "resume_version": "",
                 "cover_letter_sent": False, "confirmation": "",
                 "notes": ""}]
        out = phase6_update_tracker(apps, output_dir=tmp_path, write_file=True)
        path = out["tracker_path"]
        assert path is not None
        assert path.exists()
        wb = openpyxl.load_workbook(path)
        try:
            assert "Applications" in wb.sheetnames
            assert "Dashboard" in wb.sheetnames
        finally:
            wb.close()

    def test_skips_xlsx_when_write_file_false(self, tmp_path):
        apps = [{"company": "A", "title": "T", "score": 85, "status": "Applied",
                 "date_applied": "01/15/2026", "application_url": "",
                 "platform": "", "resume_version": "",
                 "cover_letter_sent": False, "confirmation": "",
                 "notes": ""}]
        out = phase6_update_tracker(apps, output_dir=tmp_path, write_file=False)
        # tracker_path is None when file write is disabled.
        assert out["tracker_path"] is None
        # And no .xlsx was created.
        assert list(tmp_path.glob("*.xlsx")) == []

    def test_summary_counts_statuses(self, tmp_path):
        apps = [
            {"company": "A", "title": "T1", "score": 85, "status": "Applied",
             "date_applied": "01/15/2026", "application_url": "", "platform": "",
             "resume_version": "", "cover_letter_sent": False,
             "confirmation": "", "notes": ""},
            {"company": "B", "title": "T2", "score": 70, "status": "Manual Required",
             "date_applied": "01/15/2026", "application_url": "", "platform": "",
             "resume_version": "", "cover_letter_sent": False,
             "confirmation": "", "notes": ""},
            {"company": "C", "title": "T3", "score": 50, "status": "Skipped",
             "date_applied": "01/15/2026", "application_url": "", "platform": "",
             "resume_version": "", "cover_letter_sent": False,
             "confirmation": "", "notes": ""},
        ]
        out = phase6_update_tracker(apps, output_dir=tmp_path, write_file=False)
        s = out["summary"]
        assert s["total"] == 3
        assert s["applied"] == 1
        assert s["manual"] == 1
        assert s["skipped"] == 1


# ── phase7_run_report ───────────────────────────────────────────────────────


class TestPhase7Report:
    def test_writes_md_report(self, tmp_path):
        provider = FakeProvider(report="Run Summary text.\nNext steps: ...")
        apps = [
            {"company": "Acme", "title": "FPGA", "score": 85, "status": "Applied"},
            {"company": "B", "title": "T", "score": 60, "status": "Manual Required",
             "notes": "Form requires manual review"},
            {"company": "C", "title": "T", "score": 40, "status": "Skipped"},
        ]
        out = phase7_run_report(apps, tracker_path=None, provider=provider,
                                 output_dir=tmp_path, write_file=True)
        assert "Run Summary" in out
        # File written.
        md_files = list(tmp_path.glob("*_job-application-run-report.md"))
        assert len(md_files) == 1
        text = md_files[0].read_text(encoding="utf-8")
        assert "Run Summary" in text

    def test_skips_file_when_write_file_false(self, tmp_path):
        provider = FakeProvider(report="In-page only.")
        apps = [{"company": "A", "title": "T", "score": 85, "status": "Applied"}]
        out = phase7_run_report(apps, tracker_path=None, provider=provider,
                                 output_dir=tmp_path, write_file=False)
        assert out == "In-page only."
        # No .md written.
        assert list(tmp_path.glob("*.md")) == []

    def test_top3_passed_to_provider(self):
        captured = {}

        class CapturingProvider(FakeProvider):
            def generate_report(self, summary_data):
                captured.update(summary_data)
                return "report"

        provider = CapturingProvider()
        apps = [
            {"company": "A", "title": "T", "score": 95, "status": "Applied"},
            {"company": "B", "title": "T", "score": 85, "status": "Applied"},
            {"company": "C", "title": "T", "score": 75, "status": "Applied"},
            {"company": "D", "title": "T", "score": 65, "status": "Applied"},
        ]
        phase7_run_report(apps, tracker_path=None, provider=provider,
                           output_dir=None, write_file=False)
        assert len(captured["top3_applied"]) == 3
        # Sorted by score descending.
        assert captured["top3_applied"][0][2] == 95
